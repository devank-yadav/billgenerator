import os
import io
import time
import gzip as _gz
import zipfile
import sqlite3
import secrets
import hashlib
from urllib.parse import urlencode
from datetime import datetime, date, timedelta

from flask import Flask, render_template, request, send_file, redirect, url_for, session, jsonify, Response
from functools import lru_cache

import urllib.request
import urllib.error
import json as _json

stringWidth = None

# ---------- IN-MEMORY REFERENCE DATA CACHE ----------
_ref_cache: dict = {}
_REF_TTL = 300  # seconds — 5-minute TTL, busted on write

def _cache_get(key):
    e = _ref_cache.get(key)
    return e['v'] if (e and time.time() - e['t'] < _REF_TTL) else None

def _cache_set(key, val):
    _ref_cache[key] = {'v': val, 't': time.time()}

def _cache_bust():
    _ref_cache.clear()

def _load_ref_data():
    """Return (driver_list, vehicle_rows, customer_rows) with in-process TTL caching."""
    dl = _cache_get('drivers')
    vr = _cache_get('vehicles')
    cr = _cache_get('customers')
    if dl is None or vr is None or cr is None:
        with sqlite3.connect(DATABASE) as conn:
            if dl is None:
                dl = [r[0] for r in conn.execute("SELECT name FROM drivers ORDER BY name ASC").fetchall()]
                _cache_set('drivers', dl)
            if vr is None:
                vr = list(conn.execute("SELECT name, COALESCE(vehicle_no,'') FROM vehicles ORDER BY name ASC").fetchall())
                _cache_set('vehicles', vr)
            if cr is None:
                cr = list(conn.execute("SELECT name, COALESCE(company,'') FROM customers ORDER BY name ASC").fetchall())
                _cache_set('customers', cr)
    return dl, vr, cr

GOOGLE_MAPS_API_KEY = os.environ.get('GOOGLE_MAPS_API_KEY', '')

# ---------- PDF DESIGN CONSTANTS ----------
COLOR_DARK   = (0.13, 0.13, 0.15)
COLOR_TEXT   = (0.05, 0.05, 0.07)
COLOR_MUTED  = (0.42, 0.42, 0.46)
COLOR_LINE   = (0.55, 0.55, 0.58)
COLOR_HEADER = (0.10, 0.10, 0.12)


@lru_cache(maxsize=1)
def _pdf_libs():
    # Lazy-load heavy PDF deps to reduce cold-start time.
    from reportlab.pdfgen import canvas as _canvas
    from reportlab.lib.pagesizes import A4 as _A4, landscape as _landscape
    from reportlab.pdfbase.pdfmetrics import stringWidth as _stringWidth
    from reportlab.lib.utils import ImageReader as _ImageReader
    from pypdf import PdfWriter as _PdfWriter, PdfReader as _PdfReader
    return {
        'canvas': _canvas,
        'A4': _A4,
        'landscape': _landscape,
        'stringWidth': _stringWidth,
        'ImageReader': _ImageReader,
        'PdfWriter': _PdfWriter,
        'PdfReader': _PdfReader,
    }


def _string_width(text, font_name, font_size):
    global stringWidth
    if stringWidth is None:
        stringWidth = _pdf_libs()['stringWidth']
    return stringWidth(text, font_name, font_size)


def _wrap_text(text, font_name, font_size, max_width):
    if not text:
        return []
    words = text.split()
    lines, current = [], ''
    for w in words:
        test = (current + ' ' + w).strip()
        if _string_width(test, font_name, font_size) <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = w
    if current:
        lines.append(current)
    return lines


def _format_date(d):
    try:
        return datetime.strptime(d, '%Y-%m-%d').strftime('%d / %m / %Y')
    except Exception:
        return d or ''


def _parse_ua(ua: str) -> str:
    """Return a short human-readable device label from a User-Agent string."""
    if not ua:
        return 'Unknown'
    u = ua.lower()
    if 'iphone' in u:
        return 'iPhone'
    if 'ipad' in u:
        return 'iPad'
    if 'android' in u:
        return 'Android Phone' if 'mobile' in u else 'Android Tablet'
    if 'windows' in u:
        return 'Windows'
    if 'macintosh' in u or 'mac os x' in u:
        return 'Mac'
    if 'linux' in u:
        return 'Linux'
    return 'Desktop'


def _route_positions(n_lines):
    """Return dynamic y-coordinates for route lines, driver label, and signature."""
    route_y = 248
    n = max(n_lines, 1)
    line_h = 22 if n <= 5 else 16
    line_ys = [route_y - i * line_h for i in range(n)]
    last_y = line_ys[-1]
    driver_y = max(last_y - 36, 95)
    return {'line_ys': line_ys, 'line_h': line_h, 'driver_y': driver_y}


def _draw_tracked_right(c, x_right, y, text, font_name, font_size, tracking):
    """Right-aligned text with extra letter-spacing (tracking)."""
    total_w = sum(_string_width(ch, font_name, font_size) for ch in text) + tracking * (len(text) - 1)
    x = x_right - total_w
    for ch in text:
        c.drawString(x, y, ch)
        x += _string_width(ch, font_name, font_size) + tracking


def _draw_tracked_center(c, x_center, y, text, font_name, font_size, tracking):
    total_w = sum(_string_width(ch, font_name, font_size) for ch in text) + tracking * (len(text) - 1)
    x = x_center - total_w / 2
    for ch in text:
        c.drawString(x, y, ch)
        x += _string_width(ch, font_name, font_size) + tracking


def draw_slip_template(c, width, height, n_route_lines=1):
    """Draw the static chrome of the duty slip (header, labels, lines)."""

    # ---- HEADER: contact info (top-left) ----
    c.setFont('Helvetica', 9)
    c.setFillColorRGB(*COLOR_MUTED)
    c.drawString(40, 568, 'Contact No:')
    c.drawString(40, 553, 'Address:')
    c.setFillColorRGB(*COLOR_DARK)
    c.drawString(108, 568, '9718305627, 9718305628')
    c.drawString(108, 553, 'B213, Saraswati Enclave, Gurgaon')

    # ---- HEADER: brand (top-right, two lines, wide tracking) ----
    c.setFillColorRGB(*COLOR_HEADER)
    c.setFont('Helvetica-Bold', 28)
    _draw_tracked_right(c, width - 40, 568, 'OSPREY',  'Helvetica-Bold', 28, 6)
    _draw_tracked_right(c, width - 40, 538, 'TRAVELS', 'Helvetica-Bold', 28, 6)

    # ---- Divider ----
    c.setStrokeColorRGB(*COLOR_DARK)
    c.setLineWidth(1.2)
    c.line(40, 525, width - 40, 525)

    # ---- DUTY SLIP NO + DATE block (under header, left side) ----
    c.setFillColorRGB(*COLOR_DARK)
    c.setFont('Helvetica-Bold', 10.5)
    c.drawString(40, 498, 'DUTY SLIP NO:')
    c.drawString(40, 472, 'DATE:')

    c.setStrokeColorRGB(*COLOR_LINE)
    c.setLineWidth(0.5)
    c.setDash(1, 2)
    c.line(135, 494, 320, 494)
    c.line(135, 468, 320, 468)
    c.setDash()

    # ---- TWO-COLUMN BODY ROWS ----
    rows = [
        ('COMPANY NAME:',    'CUSTOMER NAME:'),
        ('TYPE OF VEHICLE:', 'VEHICLE NO:'),
        ('STARTING KM:',     'STARTING TIME:'),
        ('CLOSING KM:',      'CLOSING TIME:'),
        ('TOTAL KM:',        'TOTAL TIME:'),
        ('PROJECT CODE:',    'MAIL APPROVAL DATE:'),
    ]
    base_y, row_h = 432, 28
    label_x_L, line_start_L, line_end_L = 40, 175, 410
    label_x_R, line_start_R, line_end_R = 445, 580, 802

    for i, (left_label, right_label) in enumerate(rows):
        y = base_y - i * row_h
        c.setFont('Helvetica-Bold', 10.5)
        c.setFillColorRGB(*COLOR_DARK)
        c.drawString(label_x_L, y, left_label)
        c.drawString(label_x_R, y, right_label)

        c.setStrokeColorRGB(*COLOR_LINE)
        c.setLineWidth(0.5)
        c.setDash(1, 2)
        c.line(line_start_L, y - 4, line_end_L, y - 4)
        c.line(line_start_R, y - 4, line_end_R, y - 4)
        c.setDash()

    # ---- ROUTE COVERED (dynamic multi-line, full width) ----
    route_y = base_y - 6 * row_h - 16  # = 248
    c.setFont('Helvetica-Bold', 10.5)
    c.setFillColorRGB(*COLOR_DARK)
    c.drawString(40, route_y, 'ROUTE COVERED:')

    pos = _route_positions(n_route_lines)
    c.setStrokeColorRGB(*COLOR_LINE)
    c.setLineWidth(0.5)
    c.setDash(1, 2)
    for ly in pos['line_ys']:
        c.line(155, ly - 4, line_end_R, ly - 4)
    c.setDash()

    # ---- DRIVER NAME (left, below route — dynamic position) ----
    driver_y = pos['driver_y']
    c.setFont('Helvetica-Bold', 10.5)
    c.setFillColorRGB(*COLOR_DARK)
    c.drawString(40, driver_y, 'DRIVER NAME:')
    c.setStrokeColorRGB(*COLOR_LINE)
    c.setLineWidth(0.5)
    c.setDash(1, 2)
    c.line(135, driver_y - 4, line_end_L, driver_y - 4)
    c.setDash()

    # ---- USER SIGNATURE (bottom right, always fixed) ----
    c.setStrokeColorRGB(*COLOR_LINE)
    c.setLineWidth(0.7)
    c.setDash(1, 2)
    c.line(575, 78, 802, 78)
    c.setDash()
    c.setFont('Helvetica-Bold', 9)
    c.setFillColorRGB(*COLOR_MUTED)
    _draw_tracked_center(c, 688, 62, 'USER SIGNATURE', 'Helvetica-Bold', 9, 1.2)


def fill_slip_data(c, data):
    """Place dynamic values into the template at the correct coordinates."""
    c.setFillColorRGB(*COLOR_TEXT)
    c.setFont('Helvetica', 11)

    # Duty slip no + Date
    c.drawString(140, 498, data.get('duty_slip_no', '') or '')
    c.drawString(140, 472, _format_date(data.get('date', '')))

    # Two-column rows
    base_y, row_h = 432, 28
    val_x_L, val_x_R = 180, 585

    left_values = [
        data.get('company_name', ''),
        data.get('vehicle_type', ''),
        data.get('starting_km', ''),
        data.get('closing_km', ''),
        data.get('total_km', ''),
        data.get('project_code', ''),
    ]
    right_values = [
        data.get('customer_name', ''),
        data.get('vehicle_no', ''),
        data.get('starting_time', ''),
        data.get('closing_time', ''),
        data.get('total_time', ''),
        _format_date(data.get('mail_approval_date', '')),
    ]
    for i, (lv, rv) in enumerate(zip(left_values, right_values)):
        y = base_y - i * row_h
        c.drawString(val_x_L, y, str(lv or ''))
        c.drawString(val_x_R, y, str(rv or ''))

    # Route covered: all wrapped lines, positions computed dynamically
    route_lines = _wrap_text(data.get('route_covered', '') or '', 'Helvetica', 11, 632)
    pos = _route_positions(len(route_lines))
    for i, line in enumerate(route_lines):
        c.drawString(160, pos['line_ys'][i], line)

    # Driver name at dynamic position
    c.drawString(140, pos['driver_y'], data.get('driver_name', '') or '')


def _build_pdf(data: dict, signature_data: str = None) -> io.BytesIO:
    import base64 as _b64
    route_lines = _wrap_text(data.get('route_covered', '') or '', 'Helvetica', 11, 632)
    n = max(len(route_lines), 1)
    pdf = _pdf_libs()
    buf = io.BytesIO()
    c = pdf['canvas'].Canvas(buf, pagesize=pdf['landscape'](pdf['A4']))
    w, h = pdf['landscape'](pdf['A4'])
    draw_slip_template(c, w, h, n_route_lines=n)
    fill_slip_data(c, data)
    # Embed digital signature above the USER SIGNATURE line (line is at y=78, label at y=62)
    if signature_data:
        try:
            raw = signature_data.split(',', 1)[1] if ',' in signature_data else signature_data
            sig_bytes = _b64.b64decode(raw)
            sig_img = pdf['ImageReader'](io.BytesIO(sig_bytes))
            # Place signature: x=576..800, bottom at y=84 (6pt above line), height=52pt
            c.drawImage(sig_img, 576, 84, width=224, height=52,
                        mask='auto', preserveAspectRatio=True, anchor='c')
        except Exception:
            pass  # Never break PDF generation over a bad signature image
    c.save()
    buf.seek(0)
    return buf


def _get_sig_for_invoice(conn, invoice_id: int):
    """Return signature_data (base64 PNG) for a signed invoice, or None."""
    rows = conn.execute(
        "SELECT invoice_ids, signature_data FROM signature_requests WHERE signed_at IS NOT NULL"
    ).fetchall()
    for (ids_json, sig_data) in rows:
        try:
            if invoice_id in _json.loads(ids_json or '[]'):
                return sig_data
        except Exception:
            pass
    return None


def _build_cover_page(title: str, slips: list, subtitle: str = '') -> io.BytesIO:
    """Portrait A4 cover/summary page prepended to batch PDF exports."""
    pdf = _pdf_libs()
    buf = io.BytesIO()
    w, h = pdf['A4']   # ~595 x 842 pt
    c = pdf['canvas'].Canvas(buf, pagesize=pdf['A4'])

    # Brand header
    c.setFillColorRGB(*COLOR_HEADER)
    c.setFont('Helvetica-Bold', 22)
    _draw_tracked_right(c, w - 40, h - 52, 'OSPREY',  'Helvetica-Bold', 22, 5)
    _draw_tracked_right(c, w - 40, h - 76, 'TRAVELS', 'Helvetica-Bold', 22, 5)
    c.setFont('Helvetica', 8)
    c.setFillColorRGB(*COLOR_MUTED)
    c.drawString(40, h - 52, '9718305627 / 9718305628')
    c.drawString(40, h - 64, 'B213, Saraswati Enclave, Gurgaon')

    c.setStrokeColorRGB(*COLOR_DARK)
    c.setLineWidth(1.0)
    c.line(40, h - 90, w - 40, h - 90)

    c.setFillColorRGB(*COLOR_DARK)
    c.setFont('Helvetica-Bold', 17)
    c.drawString(40, h - 118, title)

    y = h - 138
    if subtitle:
        c.setFont('Helvetica', 11)
        c.setFillColorRGB(*COLOR_MUTED)
        c.drawString(40, y, subtitle)
        y -= 18
    c.setFont('Helvetica', 9)
    c.setFillColorRGB(*COLOR_MUTED)
    c.drawString(40, y, f'Generated: {datetime.now().strftime("%d %b %Y, %I:%M %p")}')
    y -= 13
    c.drawString(40, y, f'Total: {len(slips)} slip{"s" if len(slips) != 1 else ""}')
    y -= 26

    # Table header
    col_x = [40, 90, 160, 262, 475]
    col_hdrs = ['SLIP #', 'DATE', 'CUSTOMER', 'ROUTE COVERED', 'STATUS']
    c.setFillColorRGB(*COLOR_DARK)
    c.rect(40, y - 4, w - 80, 18, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont('Helvetica-Bold', 8)
    for x, hdr in zip(col_x, col_hdrs):
        c.drawString(x + 3, y + 2, hdr)
    y -= 20

    c.setFont('Helvetica', 8)
    row_h = 16
    for idx, slip in enumerate(slips):
        if y < 50:
            c.setFillColorRGB(*COLOR_MUTED)
            c.drawString(44, y + 2, f'… and {len(slips) - idx} more slips')
            break
        if idx % 2 == 0:
            c.setFillColorRGB(0.96, 0.96, 0.97)
            c.rect(40, y - 2, w - 80, row_h, fill=1, stroke=0)
        c.setFillColorRGB(*COLOR_TEXT)
        route_raw = str(slip.get('route_covered', '') or '')
        route = (route_raw[:40] + '…') if len(route_raw) > 40 else route_raw
        vals = [
            str(slip.get('duty_slip_no', '') or '')[:12],
            _format_date(slip.get('date', '') or ''),
            str(slip.get('customer_name', '') or '')[:18],
            route,
            str(slip.get('bill_status', '') or ''),
        ]
        for x, val in zip(col_x, vals):
            c.drawString(x + 3, y + 3, val)
        y -= row_h

    c.save()
    buf.seek(0)
    return buf


def _build_sig_map(conn, invoice_ids: list) -> dict:
    """Return {invoice_id: signature_data} for all signed invoices in the list."""
    id_set = {int(i) for i in invoice_ids}
    rows = conn.execute(
        "SELECT invoice_ids, signature_data FROM signature_requests WHERE signed_at IS NOT NULL"
    ).fetchall()
    result = {}
    for (ids_json, sig_data) in rows:
        try:
            for inv_id in _json.loads(ids_json or '[]'):
                if inv_id in id_set and inv_id not in result:
                    result[inv_id] = sig_data
        except Exception:
            pass
    return result


app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = bool(os.environ.get('FLASK_DEBUG'))
app.secret_key = os.environ.get('SECRET_KEY') or secrets.token_hex(32)


@app.after_request
def _add_perf_headers(response):
    """Gzip text responses; add Cache-Control headers."""
    ct = response.content_type or ''
    # Long-lived cache for fingerprinted static assets
    if request.path.startswith('/static/') and response.status_code == 200:
        response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
        return response
    # No-store for authenticated HTML pages (never serve stale admin data from cache)
    if ct.startswith('text/html') and response.status_code == 200:
        response.headers.setdefault('Cache-Control', 'no-store')
    # Gzip
    if (response.status_code < 200 or response.status_code >= 300
            or 'Content-Encoding' in response.headers):
        return response
    if not (ct.startswith('text/') or 'json' in ct or 'javascript' in ct):
        return response
    if 'gzip' not in request.headers.get('Accept-Encoding', ''):
        return response
    data = response.get_data()
    if len(data) < 512:
        return response
    compressed = _gz.compress(data, compresslevel=6)
    if len(compressed) >= len(data):
        return response
    response.set_data(compressed)
    response.headers['Content-Encoding'] = 'gzip'
    response.headers['Content-Length'] = len(compressed)
    response.headers.add('Vary', 'Accept-Encoding')
    return response


def _hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode('utf-8')).hexdigest()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_DATABASE = os.path.join(BASE_DIR, 'invoices.db')
TURSO_DATABASE_URL = os.getenv('TURSO_DATABASE_URL') or os.getenv('STORAGE_URL')
TURSO_AUTH_TOKEN = os.getenv('TURSO_AUTH_TOKEN') or os.getenv('STORAGE_AUTH_TOKEN')
USE_TURSO = bool(TURSO_DATABASE_URL and TURSO_AUTH_TOKEN)


def _turso_http_url():
    url = TURSO_DATABASE_URL
    if url.startswith('libsql://'):
        url = 'https://' + url[len('libsql://'):]
    return url.rstrip('/') + '/v2/pipeline'


def _encode_arg(val):
    if val is None:
        return {"type": "null", "value": None}
    if isinstance(val, bool):
        return {"type": "integer", "value": str(int(val))}
    if isinstance(val, int):
        return {"type": "integer", "value": str(val)}
    if isinstance(val, float):
        return {"type": "float", "value": str(val)}
    return {"type": "text", "value": str(val)}


def _turso_pipeline(stmts):
    """Execute a list of (sql, args) tuples against Turso HTTP API."""
    requests_body = [
        {"type": "execute", "stmt": {"sql": sql, "args": [_encode_arg(a) for a in (args or [])]}}
        for sql, args in stmts
    ] + [{"type": "close"}]
    payload = _json.dumps({"requests": requests_body}).encode()
    req = urllib.request.Request(
        _turso_http_url(),
        data=payload,
        headers={
            "Authorization": f"Bearer {TURSO_AUTH_TOKEN}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        return _json.loads(resp.read())["results"]


def _coerce_cell(cell):
    t, v = cell.get("type"), cell.get("value")
    if t == "null" or v is None:
        return None
    if t == "integer":
        try:
            return int(v)
        except (ValueError, TypeError):
            return v
    if t == "float":
        try:
            return float(v)
        except (ValueError, TypeError):
            return v
    return v  # text / blob → keep as string


class _TursoCursor:
    def __init__(self, result):
        r = result.get("response", {}).get("result", {})
        cols = [c["name"] for c in r.get("cols", [])]
        self.lastrowid = r.get("last_insert_rowid")
        if self.lastrowid is not None:
            try:
                self.lastrowid = int(self.lastrowid)
            except (ValueError, TypeError):
                pass
        raw = r.get("rows", [])
        self._rows = []
        for raw_row in raw:
            self._rows.append(tuple(_coerce_cell(cell) for cell in raw_row))
        self._idx = 0

    def fetchone(self):
        if self._idx < len(self._rows):
            row = self._rows[self._idx]
            self._idx += 1
            return row
        return None

    def fetchall(self):
        return self._rows[self._idx:]

    def __iter__(self):
        return iter(self._rows)


class _TursoConnection:
    def __init__(self):
        self._stmts = []

    def execute(self, sql, parameters=()):
        results = _turso_pipeline([(sql, list(parameters))])
        if results[0].get("type") == "error":
            raise sqlite3.OperationalError(results[0]["error"]["message"])
        return _TursoCursor(results[0])

    def executemany(self, sql, seq_of_params):
        stmts = [(sql, list(p)) for p in seq_of_params]
        _turso_pipeline(stmts)

    def commit(self):
        pass

    def close(self):
        pass

    def __enter__(self):
        return self

    def __exit__(self, *_):
        pass


_sqlite_connect = sqlite3.connect


def _db_connect(database, *args, **kwargs):
    if USE_TURSO:
        return _TursoConnection()
    return _sqlite_connect(database, *args, **kwargs)


# Keep existing sqlite3.connect(...) call sites working with Turso.
sqlite3.connect = _db_connect


class _MatCursor:
    """Materialised cursor — rows fetched eagerly so connection can close."""
    def __init__(self, rows):
        self._rows = rows
        self._idx = 0

    def fetchone(self):
        if self._idx < len(self._rows):
            r = self._rows[self._idx]; self._idx += 1; return r
        return None

    def fetchall(self):
        return self._rows[self._idx:]


def _db_multi_exec(queries):
    """Run [(sql, params), ...] returning a list of cursor-like objects.
    On Turso: one HTTP round-trip for all queries.
    On SQLite: sequential execution, rows materialised before connection closes."""
    if USE_TURSO:
        results = _turso_pipeline([(sql, list(params)) for sql, params in queries])
        return [_TursoCursor(r) for r in results[:-1]]
    conn = _sqlite_connect(DATABASE)
    try:
        cursors = [_MatCursor(conn.execute(sql, params).fetchall()) for sql, params in queries]
    finally:
        conn.close()
    return cursors

if USE_TURSO:
    DATABASE = TURSO_DATABASE_URL  # value unused for Turso path but kept for consistency
elif os.getenv('VERCEL'):
    DATABASE = '/tmp/invoices.db'
else:
    DATABASE = SOURCE_DATABASE


def init_db():
    with sqlite3.connect(DATABASE) as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_name TEXT,
                date TEXT,
                file_path TEXT,
                admin_username TEXT
            )
        ''')
        columns = [col[1] for col in conn.execute("PRAGMA table_info(invoices)").fetchall()]
        required_columns = [
            ('created_at', 'TEXT'),
            ('driver_name', 'TEXT'),
            ('company_name', 'TEXT'),
            ('duty_slip_no', 'TEXT'),
            ('vehicle_type', 'TEXT'),
            ('vehicle_no', 'TEXT'),
            ('starting_km', 'TEXT'),
            ('closing_km', 'TEXT'),
            ('total_km', 'TEXT'),
            ('starting_time', 'TEXT'),
            ('closing_time', 'TEXT'),
            ('total_time', 'TEXT'),
            ('dn', 'TEXT'),
            ('remarks', 'TEXT'),
            ('route_covered', 'TEXT'),
            ('bill_status', 'TEXT'),
            ('payment_date', 'TEXT'),
            ('project_code', 'TEXT'),
            ('mail_approval_date', 'TEXT'),
        ]
        for col_name, col_type in required_columns:
            if col_name not in columns:
                conn.execute(f"ALTER TABLE invoices ADD COLUMN {col_name} {col_type}")
        # Backfill status for existing rows
        conn.execute("UPDATE invoices SET bill_status = 'Bill Generated' WHERE bill_status IS NULL")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_date ON invoices(date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_created_at ON invoices(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_bill_status ON invoices(bill_status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_driver_name ON invoices(driver_name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_vehicle_type ON invoices(vehicle_type)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_customer_name ON invoices(customer_name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_duty_slip_no ON invoices(duty_slip_no)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invoices_signature_status ON invoices(signature_status)")

        conn.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE,
                password TEXT
            )
        ''')
        existing = conn.execute("SELECT id, password FROM users WHERE username = ?", ("admin",)).fetchone()
        if not existing:
            conn.execute("INSERT INTO users (username, password) VALUES (?, ?)", ("admin", _hash_pw("admin")))
        elif existing[1] and len(existing[1]) != 64:
            # Migrate plaintext password to sha256 hash
            conn.execute("UPDATE users SET password = ? WHERE id = ?", (_hash_pw(existing[1]), existing[0]))

        conn.execute('''
            CREATE TABLE IF NOT EXISTS drivers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS config (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS vehicles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                vehicle_no TEXT
            )
        ''')
        # Migrate vehicles table — add vehicle_no if missing
        veh_cols = [col[1] for col in conn.execute("PRAGMA table_info(vehicles)").fetchall()]
        if 'vehicle_no' not in veh_cols:
            conn.execute("ALTER TABLE vehicles ADD COLUMN vehicle_no TEXT")

        conn.execute('''
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                company TEXT
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS slip_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                admin_username TEXT NOT NULL,
                template_name TEXT NOT NULL,
                customer_name TEXT, company_name TEXT,
                vehicle_type TEXT, vehicle_no TEXT,
                route_covered TEXT, dn TEXT, remarks TEXT,
                driver_name TEXT, starting_km TEXT, total_km TEXT,
                created_at TEXT,
                use_count INTEGER DEFAULT 0,
                last_used TEXT
            )
        ''')
        # Migrate existing slip_templates — add new columns if missing
        tpl_cols = [col[1] for col in conn.execute("PRAGMA table_info(slip_templates)").fetchall()]
        if 'use_count' not in tpl_cols:
            conn.execute("ALTER TABLE slip_templates ADD COLUMN use_count INTEGER DEFAULT 0")
        if 'last_used' not in tpl_cols:
            conn.execute("ALTER TABLE slip_templates ADD COLUMN last_used TEXT")
        if 'project_code' not in tpl_cols:
            conn.execute("ALTER TABLE slip_templates ADD COLUMN project_code TEXT")
        if 'mail_approval_date' not in tpl_cols:
            conn.execute("ALTER TABLE slip_templates ADD COLUMN mail_approval_date TEXT")
        if 'starting_time' not in tpl_cols:
            conn.execute("ALTER TABLE slip_templates ADD COLUMN starting_time TEXT")
        if 'closing_time' not in tpl_cols:
            conn.execute("ALTER TABLE slip_templates ADD COLUMN closing_time TEXT")
        if 'route_stops_json' not in tpl_cols:
            conn.execute("ALTER TABLE slip_templates ADD COLUMN route_stops_json TEXT")

        # Signature columns on invoices
        inv_cols = [col[1] for col in conn.execute("PRAGMA table_info(invoices)").fetchall()]
        if 'signature_status' not in inv_cols:
            conn.execute("ALTER TABLE invoices ADD COLUMN signature_status TEXT")
        if 'signed_at' not in inv_cols:
            conn.execute("ALTER TABLE invoices ADD COLUMN signed_at TEXT")

        conn.execute('''
            CREATE TABLE IF NOT EXISTS signature_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                token TEXT UNIQUE NOT NULL,
                invoice_ids TEXT NOT NULL,
                customer_name TEXT,
                created_at TEXT,
                expires_at TEXT,
                signed_at TEXT,
                signature_data TEXT,
                signer_ip TEXT,
                signer_ua TEXT
            )
        ''')
        sig_cols = [col[1] for col in conn.execute("PRAGMA table_info(signature_requests)").fetchall()]
        if 'signer_ua' not in sig_cols:
            conn.execute("ALTER TABLE signature_requests ADD COLUMN signer_ua TEXT")

        # Customer portal token
        cust_cols = [col[1] for col in conn.execute("PRAGMA table_info(customers)").fetchall()]
        if 'portal_token' not in cust_cols:
            conn.execute("ALTER TABLE customers ADD COLUMN portal_token TEXT")

        conn.execute('''
            CREATE TABLE IF NOT EXISTS recurring_trips (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                customer_name TEXT,
                company_name TEXT,
                vehicle_type TEXT,
                vehicle_no TEXT,
                driver_name TEXT,
                route_covered TEXT,
                project_code TEXT,
                days_of_week TEXT,
                active INTEGER DEFAULT 1,
                last_generated TEXT,
                created_at TEXT
            )
        ''')


init_db()


def delete_invoice_file(invoice_id):
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("DELETE FROM invoices WHERE id = ?", (invoice_id,))


def _canonical_name(conn, table, name):
    """Return the already-stored canonical name (preserving its original case) if a
    case-insensitive match exists, otherwise return name as-is."""
    row = conn.execute(
        f"SELECT name FROM {table} WHERE LOWER(name) = LOWER(?)", (name,)
    ).fetchone()
    return row[0] if row else name


def get_next_duty_slip_no():
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(
            "SELECT duty_slip_no FROM invoices WHERE duty_slip_no IS NOT NULL AND duty_slip_no != '' ORDER BY id DESC LIMIT 200",
        ).fetchall()
    max_num = 0
    for (s,) in rows:
        digits = ''.join(c for c in s if c.isdigit())
        if digits:
            n = int(digits)
            if n > max_num:
                max_num = n
    return str(max_num + 1) if max_num > 0 else ''


_FAVICON_SVG = (
    b'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32">'
    b'<rect width="32" height="32" rx="7" fill="#007aff"/>'
    b'<text x="16" y="22" text-anchor="middle" font-family="-apple-system,system-ui,sans-serif"'
    b' font-weight="900" font-size="15" fill="white">OT</text></svg>'
)


@app.route('/favicon.ico')
@app.route('/favicon.svg')
def favicon():
    return Response(_FAVICON_SVG, mimetype='image/svg+xml',
                    headers={'Cache-Control': 'public, max-age=604800'})


@app.route('/')
def home():
    if 'admin' in session:
        return redirect(url_for('admin_portal'))
    return redirect(url_for('login_page'))


@app.route('/login')
def login_page():
    return render_template('admin_login.html')


@app.route('/admin_login', methods=['POST'])
def admin_login():
    username = request.form['username']
    password = request.form['password']
    with sqlite3.connect(DATABASE) as conn:
        user = conn.execute(
            "SELECT * FROM users WHERE username = ? AND password = ?",
            (username, _hash_pw(password))
        ).fetchone()
    if user:
        session['admin'] = username
        return redirect(url_for('admin_portal'))
    return render_template('admin_login.html', error="Invalid credentials")


@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('home'))


# --------------------------
# ADMIN PORTAL
# --------------------------
@app.route('/admin_portal')
def admin_portal(_partial=False):
    if 'admin' not in session:
        return redirect(url_for('home'))

    search_q       = request.args.get('q', '').strip()
    customer_name  = request.args.get('customer_name', '').strip()
    driver_filter  = request.args.get('driver', '').strip()
    vehicle_filter = request.args.get('vehicle', '').strip()
    month_filter   = request.args.get('month', '')
    date_from      = request.args.get('date_from', '')
    date_to        = request.args.get('date_to', '')
    status_filter  = request.args.get('status', '')
    date_quick     = request.args.get('date_quick', '')    # today | week | month | last_month | last7
    status_quick   = request.args.get('status_quick', '')  # pending_bill | pending_pay | paid
    sig_filter     = request.args.get('sig_filter', '')    # none | pending | signed

    page = max(int(request.args.get('page', 1) or 1), 1)
    page_size = int(request.args.get('page_size', 10) or 10)
    page_size = min(max(page_size, 10), 200)
    offset = (page - 1) * page_size
    base_args = request.args.to_dict(flat=True)
    base_args.pop('page', None)
    base_args.pop('page_size', None)
    base_query = urlencode({k: v for k, v in base_args.items() if v not in (None, '')})

    where_clause = "WHERE 1=1"
    params = []

    # Date quick filter
    today = date.today()
    if date_quick == 'today':
        where_clause += " AND date = ?"
        params.append(today.strftime('%Y-%m-%d'))
    elif date_quick == 'week':
        monday = today - timedelta(days=today.weekday())
        where_clause += " AND date >= ?"
        params.append(monday.strftime('%Y-%m-%d'))
    elif date_quick == 'month':
        where_clause += " AND strftime('%Y-%m', date) = ?"
        params.append(today.strftime('%Y-%m'))
    elif date_quick == 'last_month':
        first_this = today.replace(day=1)
        last_prev  = first_this - timedelta(days=1)
        where_clause += " AND strftime('%Y-%m', date) = ?"
        params.append(last_prev.strftime('%Y-%m'))
    elif date_quick == 'last7':
        seven_ago = today - timedelta(days=7)
        where_clause += " AND date >= ?"
        params.append(seven_ago.strftime('%Y-%m-%d'))

    # Status quick filter (independent — can combine with date_quick)
    if status_quick == 'pending_bill':
        where_clause += " AND COALESCE(bill_status,'Bill Generated') = 'Bill Generated'"
    elif status_quick == 'pending_pay':
        where_clause += " AND bill_status = 'Bill Submitted'"
    elif status_quick == 'paid':
        where_clause += " AND bill_status = 'Payment Received'"

    # Signature filter
    if sig_filter == 'none':
        where_clause += " AND COALESCE(signature_status, '') = ''"
    elif sig_filter == 'pending':
        where_clause += " AND signature_status = 'pending'"
    elif sig_filter == 'signed':
        where_clause += " AND signature_status = 'signed'"

    if search_q:
        where_clause += " AND (customer_name LIKE ? OR company_name LIKE ? OR vehicle_no LIKE ? OR vehicle_type LIKE ? OR driver_name LIKE ? OR duty_slip_no LIKE ?)"
        like = f"%{search_q}%"
        params.extend([like, like, like, like, like, like])
    if customer_name:
        where_clause += " AND customer_name LIKE ?"
        params.append(f"%{customer_name}%")
    if driver_filter:
        where_clause += " AND driver_name = ?"
        params.append(driver_filter)
    if vehicle_filter:
        where_clause += " AND vehicle_type = ?"
        params.append(vehicle_filter)
    if month_filter:
        where_clause += " AND strftime('%Y-%m', created_at) = ?"
        params.append(month_filter)
    if date_from:
        where_clause += " AND date >= ?"
        params.append(date_from)
    if date_to:
        where_clause += " AND date <= ?"
        params.append(date_to)
    if status_filter:
        where_clause += " AND bill_status = ?"
        params.append(status_filter)

    query = f"""
        SELECT id, customer_name, company_name, date, duty_slip_no,
               file_path, created_at, driver_name,
               COALESCE(bill_status, 'Bill Generated') as bill_status,
               payment_date,
               vehicle_type, vehicle_no, starting_km, closing_km, total_km,
               starting_time, closing_time, total_time, dn, remarks, route_covered,
               project_code, mail_approval_date,
               COALESCE(signature_status, '') as signature_status
        FROM invoices {where_clause}
        ORDER BY id DESC
        LIMIT ? OFFSET ?
    """

    _ym = today.strftime('%Y-%m')
    _stats_sql = """
        SELECT COUNT(*) AS total,
          SUM(CASE WHEN strftime('%Y-%m', date) = ? THEN 1 ELSE 0 END) AS this_month,
          SUM(CASE WHEN COALESCE(bill_status,'Bill Generated') = 'Bill Generated' THEN 1 ELSE 0 END) AS pending_bill,
          SUM(CASE WHEN bill_status = 'Bill Submitted' THEN 1 ELSE 0 END) AS pending_pay,
          SUM(CASE WHEN bill_status = 'Payment Received' AND strftime('%Y-%m', payment_date) = ? THEN 1 ELSE 0 END) AS paid_month
        FROM invoices
    """
    # Batch all 5 queries → 1 HTTP round-trip on Turso
    _curs = _db_multi_exec([
        (f"SELECT COUNT(*) FROM invoices {where_clause}", tuple(params)),
        (query, tuple(params + [page_size, offset])),
        ("SELECT DISTINCT driver_name FROM invoices WHERE driver_name IS NOT NULL AND driver_name != '' ORDER BY driver_name ASC", ()),
        ("SELECT DISTINCT vehicle_type FROM invoices WHERE vehicle_type IS NOT NULL AND vehicle_type != '' ORDER BY vehicle_type ASC", ()),
        (_stats_sql, (_ym, _ym)),
    ])
    total_results  = (_curs[0].fetchone() or (0,))[0]
    invoices       = _curs[1].fetchall()
    driver_names   = [r[0] for r in _curs[2].fetchall() if r[0]]
    vehicle_names  = [r[0] for r in _curs[3].fetchall() if r[0]]
    stats          = _curs[4].fetchone()

    stats_total = stats[0] if stats else 0
    stats_this_month = stats[1] if stats else 0
    stats_pending_bill = stats[2] if stats else 0
    stats_pending_pay = stats[3] if stats else 0
    stats_paid_month = stats[4] if stats else 0

    ctx = dict(
        invoices=invoices,
        driver_names=driver_names,
        vehicle_names=vehicle_names,
        search_q=search_q,
        customer_name=customer_name,
        driver_filter=driver_filter,
        vehicle_filter=vehicle_filter,
        month_filter=month_filter,
        date_from=date_from,
        date_to=date_to,
        status_filter=status_filter,
        date_quick=date_quick,
        status_quick=status_quick,
        sig_filter=sig_filter,
        stats_total=stats_total,
        stats_this_month=stats_this_month,
        stats_pending_bill=stats_pending_bill,
        stats_pending_pay=stats_pending_pay,
        stats_paid_month=stats_paid_month,
        today_str=today.strftime('%A, %d %B %Y'),
        page=page,
        page_size=page_size,
        total_results=total_results,
        base_query=base_query,
    )
    if _partial:
        return render_template('_slips_partial.html', **ctx)
    return render_template('admin_portal.html', **ctx)


@app.route('/admin/portal/slips')
def portal_slips_fragment():
    """Return only the slips table region as an HTML partial for AJAX pagination."""
    if 'admin' not in session:
        return '', 401
    # Reuse same filter/pagination logic as admin_portal but render partial only
    return admin_portal(_partial=True)


@app.route('/update_status/<int:invoice_id>', methods=['POST'])
def update_status(invoice_id):
    if 'admin' not in session:
        return jsonify({'ok': False}), 401
    new_status = request.form.get('status', 'Bill Generated')
    payment_date = request.form.get('payment_date', '') or None
    if new_status != 'Payment Received':
        payment_date = None
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            "UPDATE invoices SET bill_status = ?, payment_date = ? WHERE id = ?",
            (new_status, payment_date, invoice_id)
        )
    return jsonify({'ok': True})


@app.route('/bulk_update_status', methods=['POST'])
def bulk_update_status():
    if 'admin' not in session:
        return jsonify({'ok': False}), 401
    selected_ids = request.form.getlist('selected_invoices')
    status = request.form.get('status', 'Bill Generated')
    payment_date = request.form.get('payment_date', '') or None
    if status != 'Payment Received':
        payment_date = None
    if not selected_ids:
        return jsonify({'ok': False, 'error': 'No invoices selected'})
    with sqlite3.connect(DATABASE) as conn:
        placeholders = ','.join('?' * len(selected_ids))
        conn.execute(
            f"UPDATE invoices SET bill_status = ?, payment_date = ? WHERE id IN ({placeholders})",
            [status, payment_date] + selected_ids
        )
    return jsonify({'ok': True, 'updated': len(selected_ids), 'status': status, 'payment_date': payment_date or ''})


@app.route('/delete_invoice/<int:invoice_id>', methods=['POST'])
def delete_invoice(invoice_id):
    if 'admin' not in session:
        return redirect(url_for('home'))
    delete_invoice_file(invoice_id)
    return redirect(url_for('admin_portal'))


@app.route('/bulk_action', methods=['POST'])
def bulk_action():
    if 'admin' not in session:
        return redirect(url_for('home'))

    selected_ids = request.form.getlist('selected_invoices')
    action = request.form.get('action')
    if not selected_ids:
        return redirect(url_for('admin_portal'))

    with sqlite3.connect(DATABASE) as conn:
        placeholders = ','.join('?' for _ in selected_ids)

        if action == 'delete':
            conn.execute(f"DELETE FROM invoices WHERE id IN ({placeholders})", selected_ids)
            return redirect(url_for('admin_portal'))

        elif action == 'download':
            rows = conn.execute(
                f"""SELECT id, duty_slip_no, customer_name, company_name, date,
                           vehicle_type, vehicle_no, starting_km, closing_km, total_km,
                           starting_time, closing_time, total_time,
                           project_code, mail_approval_date, route_covered, driver_name
                    FROM invoices WHERE id IN ({placeholders})""",
                selected_ids
            ).fetchall()
            sig_map = _build_sig_map(conn, [int(i) for i in selected_ids])
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zipf:
                for r in rows:
                    data = {
                        'duty_slip_no': r[1], 'customer_name': r[2], 'company_name': r[3],
                        'date': r[4], 'vehicle_type': r[5], 'vehicle_no': r[6],
                        'starting_km': r[7], 'closing_km': r[8], 'total_km': r[9],
                        'starting_time': r[10], 'closing_time': r[11], 'total_time': r[12],
                        'project_code': r[13], 'mail_approval_date': r[14],
                        'route_covered': r[15], 'driver_name': r[16],
                    }
                    pdf_buf = _build_pdf(data, signature_data=sig_map.get(r[0]))
                    fname = f"slip_{r[1] or r[2].replace(' ', '_')}.pdf"
                    zipf.writestr(fname, pdf_buf.read())
            zip_buffer.seek(0)
            return send_file(zip_buffer, as_attachment=True, download_name='invoices.zip', mimetype='application/zip')

        elif action == 'excel':
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            rows = conn.execute(
                f"""SELECT id, duty_slip_no, date, customer_name, company_name,
                           driver_name, vehicle_type, vehicle_no,
                           starting_km, closing_km, total_km,
                           starting_time, closing_time, total_time,
                           route_covered, dn, remarks, project_code,
                           mail_approval_date, bill_status, payment_date,
                           created_at
                    FROM invoices WHERE id IN ({placeholders})
                    ORDER BY date DESC, id DESC""",
                selected_ids,
            ).fetchall()
            wb = Workbook()
            ws = wb.active
            ws.title = "Duty Slips"
            headers = ["ID", "Slip #", "Date", "Customer", "Company",
                       "Driver", "Vehicle Type", "Vehicle No",
                       "Start KM", "Close KM", "Total KM",
                       "Start Time", "Close Time", "Total Time",
                       "Route", "DN", "Remarks", "Project Code",
                       "Mail Approval Date", "Status", "Payment Date",
                       "Created At"]
            ws.append(headers)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="007AFF")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in rows:
                ws.append(["" if v is None else v for v in row])
            for col_idx, h in enumerate(headers, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(h) + 2)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            return send_file(buf, as_attachment=True,
                             download_name=f'duty_slips_{stamp}.xlsx',
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        elif action == 'print':
            rows = conn.execute(
                f"""SELECT id, duty_slip_no, customer_name, company_name, date,
                           vehicle_type, vehicle_no, starting_km, closing_km, total_km,
                           starting_time, closing_time, total_time,
                           project_code, mail_approval_date, route_covered, driver_name,
                           COALESCE(bill_status, 'Bill Generated')
                    FROM invoices WHERE id IN ({placeholders})""",
                selected_ids
            ).fetchall()
            sig_map = _build_sig_map(conn, [int(i) for i in selected_ids])
            pdf = _pdf_libs()
            writer = pdf['PdfWriter']()
            # Cover page
            slip_summaries = [
                {'duty_slip_no': r[1], 'customer_name': r[2], 'date': r[4],
                 'route_covered': r[15], 'bill_status': r[17]}
                for r in rows
            ]
            customers_in_batch = list(dict.fromkeys(r[2] for r in rows if r[2]))
            cover_subtitle = customers_in_batch[0] if len(customers_in_batch) == 1 else f'{len(customers_in_batch)} customers'
            cover_reader = pdf['PdfReader'](io.BytesIO(_build_cover_page('Batch Export', slip_summaries, cover_subtitle).read()))
            for page in cover_reader.pages:
                writer.add_page(page)
            for r in rows:
                data = {
                    'duty_slip_no': r[1], 'customer_name': r[2], 'company_name': r[3],
                    'date': r[4], 'vehicle_type': r[5], 'vehicle_no': r[6],
                    'starting_km': r[7], 'closing_km': r[8], 'total_km': r[9],
                    'starting_time': r[10], 'closing_time': r[11], 'total_time': r[12],
                    'project_code': r[13], 'mail_approval_date': r[14],
                    'route_covered': r[15], 'driver_name': r[16],
                }
                reader = pdf['PdfReader'](io.BytesIO(_build_pdf(data, signature_data=sig_map.get(r[0])).read()))
                for page in reader.pages:
                    writer.add_page(page)
            merged_buffer = io.BytesIO()
            writer.write(merged_buffer)
            merged_buffer.seek(0)
            return send_file(merged_buffer, mimetype='application/pdf', download_name='print_batch.pdf', as_attachment=False)

    return redirect(url_for('admin_portal'))


# --------------------------
# GENERATOR
# --------------------------
@app.route('/generator')
def generator_form():
    if 'admin' not in session:
        return redirect(url_for('home'))
    prefill = {}
    template_id = request.args.get('template_id', type=int)
    driver_list, vehicle_rows, customer_rows = _load_ref_data()
    with sqlite3.connect(DATABASE) as conn:
        quick_templates = conn.execute(
            "SELECT id, template_name FROM slip_templates WHERE admin_username = ? ORDER BY use_count DESC, created_at DESC",
            (session['admin'],)
        ).fetchall()
        if template_id:
            t = conn.execute(
                """SELECT customer_name, company_name, vehicle_type, vehicle_no,
                          route_covered, dn, remarks, driver_name,
                          starting_km, total_km,
                          COALESCE(project_code,''), COALESCE(mail_approval_date,''),
                          COALESCE(starting_time,''), COALESCE(closing_time,''),
                          COALESCE(route_stops_json,'')
                   FROM slip_templates
                   WHERE id = ? AND admin_username = ?""",
                (template_id, session['admin']),
            ).fetchone()
            if t:
                prefill = {
                    'customer_name': t[0] or '', 'company_name': t[1] or '',
                    'vehicle_type':  t[2] or '', 'vehicle_no':   t[3] or '',
                    'route_covered': t[4] or '', 'dn':           t[5] or '',
                    'remarks':       t[6] or '', 'driver_name':  t[7] or '',
                    'starting_km':   t[8] or '', 'total_km':     t[9] or '',
                    'project_code':  t[10] or '', 'mail_approval_date': t[11] or '',
                    'starting_time': t[12] or '', 'closing_time': t[13] or '',
                    'route_stops_json': t[14] or '',
                }
    vehicle_list  = [r[0] for r in vehicle_rows]
    vehicle_map   = {r[0]: r[1] for r in vehicle_rows}
    customer_list = [r[0] for r in customer_rows]
    customer_map  = {r[0]: r[1] for r in customer_rows}
    next_slip_no = get_next_duty_slip_no()
    today = date.today().strftime('%Y-%m-%d')
    return render_template('generator.html',
                           driver_list=driver_list, vehicle_list=vehicle_list, vehicle_map=vehicle_map,
                           customer_list=customer_list, customer_map=customer_map,
                           quick_templates=quick_templates,
                           prefill=prefill, next_slip_no=next_slip_no, today=today)


@app.route('/clone/<int:invoice_id>')
def clone_invoice(invoice_id):
    if 'admin' not in session:
        return redirect(url_for('home'))
    driver_list, vehicle_rows, customer_rows = _load_ref_data()
    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            "SELECT customer_name, company_name, vehicle_type, vehicle_no, driver_name, project_code, mail_approval_date, route_covered FROM invoices WHERE id = ?",
            (invoice_id,)
        ).fetchone()
    if not row:
        return redirect(url_for('generator_form'))
    vehicle_list  = [r[0] for r in vehicle_rows]
    vehicle_map   = {r[0]: r[1] for r in vehicle_rows}
    customer_list = [r[0] for r in customer_rows]
    customer_map  = {r[0]: r[1] for r in customer_rows}
    prefill = {
        'customer_name': row[0] or '',
        'company_name': row[1] or '',
        'vehicle_type': row[2] or '',
        'vehicle_no': row[3] or '',
        'driver_name': row[4] or '',
        'project_code': row[5] or '',
        'mail_approval_date': row[6] or '',
        'route_covered': row[7] or '',
    }
    next_slip_no = get_next_duty_slip_no()
    today = date.today().strftime('%Y-%m-%d')
    return render_template('generator.html',
                           driver_list=driver_list, vehicle_list=vehicle_list, vehicle_map=vehicle_map,
                           customer_list=customer_list, customer_map=customer_map,
                           quick_templates=[],
                           prefill=prefill, next_slip_no=next_slip_no, today=today)


@app.route('/generate', methods=['POST'])
def generate_invoice():
    if 'admin' not in session:
        return redirect(url_for('home'))

    admin_username = session['admin']
    customer_name = request.form['customer_name']
    company_name = request.form['company_name']
    date_value = request.form['date']
    duty_slip_no = request.form['duty_slip_no']
    vehicle_type = request.form['vehicle_type']
    vehicle_no = request.form['vehicle_no']
    starting_km = request.form['starting_km']
    closing_km = request.form['closing_km']
    total_km = request.form['total_km']
    starting_time = request.form['starting_time']
    closing_time = request.form['closing_time']
    total_time = request.form['total_time']
    project_code = request.form.get('project_code', '')
    mail_approval_date = request.form.get('mail_approval_date', '')
    route_covered = request.form['route_covered']
    driver_name = request.form['driver_name']

    # Normalize names to existing canonical versions to prevent case-duplicate records
    with sqlite3.connect(DATABASE) as _norm_conn:
        if customer_name:
            customer_name = _canonical_name(_norm_conn, 'customers', customer_name)
        if vehicle_type:
            vehicle_type = _canonical_name(_norm_conn, 'vehicles', vehicle_type)

    slip_data = {
        'customer_name': customer_name,
        'company_name': company_name,
        'date': date_value,
        'duty_slip_no': duty_slip_no,
        'vehicle_type': vehicle_type,
        'vehicle_no': vehicle_no,
        'starting_km': starting_km,
        'closing_km': closing_km,
        'total_km': total_km,
        'starting_time': starting_time,
        'closing_time': closing_time,
        'total_time': total_time,
        'project_code': project_code,
        'mail_approval_date': mail_approval_date,
        'route_covered': route_covered,
        'driver_name': driver_name,
    }
    buf = _build_pdf(slip_data)

    created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("""
            INSERT INTO invoices(
                customer_name, company_name, date,
                duty_slip_no, vehicle_type, vehicle_no,
                starting_km, closing_km, total_km,
                starting_time, closing_time, total_time,
                project_code, mail_approval_date, route_covered, driver_name,
                admin_username, created_at, bill_status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Bill Generated')
        """, (
            customer_name, company_name, date_value,
            duty_slip_no, vehicle_type, vehicle_no,
            starting_km, closing_km, total_km,
            starting_time, closing_time, total_time,
            project_code, mail_approval_date, route_covered, driver_name,
            admin_username, created_at
        ))
        # Auto-create customer if not already in the list
        if customer_name:
            exists = conn.execute(
                "SELECT 1 FROM customers WHERE LOWER(name) = LOWER(?)", (customer_name,)
            ).fetchone()
            if not exists:
                conn.execute(
                    "INSERT INTO customers (name, company) VALUES (?, ?)",
                    (customer_name, company_name or '')
                )
        # Auto-create vehicle if not already in the list
        if vehicle_type:
            exists = conn.execute(
                "SELECT 1 FROM vehicles WHERE LOWER(name) = LOWER(?)", (vehicle_type,)
            ).fetchone()
            if not exists:
                conn.execute(
                    "INSERT INTO vehicles (name, vehicle_no) VALUES (?, ?)",
                    (vehicle_type, vehicle_no or '')
                )

    _cache_bust()
    filename = f"slip_{duty_slip_no or customer_name.replace(' ', '_')}.pdf"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


@app.route('/templates/<int:template_id>/json', methods=['GET'])
def template_json(template_id):
    if 'admin' not in session:
        return jsonify({}), 401
    with sqlite3.connect(DATABASE) as conn:
        t = conn.execute(
            """SELECT template_name, customer_name, company_name, vehicle_type, vehicle_no,
                      route_covered, driver_name, starting_km, total_km,
                      COALESCE(project_code,''), COALESCE(mail_approval_date,''),
                      COALESCE(starting_time,''), COALESCE(closing_time,''),
                      COALESCE(route_stops_json,'')
               FROM slip_templates WHERE id = ? AND admin_username = ?""",
            (template_id, session['admin'])
        ).fetchone()
    if not t:
        return jsonify({}), 404
    return jsonify({
        'template_name': t[0], 'customer_name': t[1], 'company_name': t[2],
        'vehicle_type': t[3], 'vehicle_no': t[4], 'route_covered': t[5],
        'driver_name': t[6], 'starting_km': t[7], 'total_km': t[8],
        'project_code': t[9], 'mail_approval_date': t[10],
        'starting_time': t[11], 'closing_time': t[12], 'route_stops_json': t[13],
    })


@app.route('/invoice/<int:invoice_id>/download')
def download_invoice(invoice_id):
    if 'admin' not in session:
        return redirect(url_for('login_page'))
    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            """SELECT customer_name, company_name, date, duty_slip_no, vehicle_type,
                      vehicle_no, starting_km, closing_km, total_km,
                      starting_time, closing_time, total_time,
                      project_code, mail_approval_date, route_covered, driver_name
               FROM invoices WHERE id = ? AND admin_username = ?""",
            (invoice_id, session['admin'])
        ).fetchone()
        if not row:
            return "Invoice not found", 404
        sig_data = _get_sig_for_invoice(conn, invoice_id)
    data = {
        'customer_name': row[0], 'company_name': row[1], 'date': row[2],
        'duty_slip_no': row[3], 'vehicle_type': row[4], 'vehicle_no': row[5],
        'starting_km': row[6], 'closing_km': row[7], 'total_km': row[8],
        'starting_time': row[9], 'closing_time': row[10], 'total_time': row[11],
        'project_code': row[12], 'mail_approval_date': row[13],
        'route_covered': row[14], 'driver_name': row[15],
    }
    buf = _build_pdf(data, signature_data=sig_data)
    filename = f"slip_{row[3] or row[0].replace(' ', '_')}.pdf"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


@app.route('/config/maps_key', methods=['GET'])
def get_maps_key():
    if 'admin' not in session:
        return jsonify({'key': ''}), 401
    return jsonify({'key': GOOGLE_MAPS_API_KEY})


@app.route('/customer_autocomplete')
def customer_autocomplete():
    if 'admin' not in session:
        return jsonify([])
    query = request.args.get('query', '').strip()
    if not query:
        return jsonify([])
    with sqlite3.connect(DATABASE) as conn:
        results = [row[0] for row in conn.execute("""
            SELECT DISTINCT customer_name FROM invoices
            WHERE customer_name LIKE ?
            ORDER BY customer_name ASC LIMIT 10
        """, (f"%{query}%",)).fetchall()]
    return jsonify(results)


# --------------------------
# SETTINGS
# --------------------------
@app.route('/settings', methods=['GET', 'POST'])
def settings_page():
    if 'admin' not in session:
        return redirect(url_for('home'))

    message = ''
    message_type = 'success'

    if request.method == 'POST':
        action = request.form.get('action', 'add_driver')
        if action == 'add_driver':
            new_driver = request.form.get('new_driver', '').strip()
            if new_driver:
                with sqlite3.connect(DATABASE) as conn:
                    try:
                        conn.execute("INSERT INTO drivers (name) VALUES (?)", (new_driver,))
                        message = f"Driver '{new_driver}' added successfully."
                    except sqlite3.IntegrityError:
                        message = f"Driver '{new_driver}' already exists."
                        message_type = 'error'
        elif action == 'delete_driver':
            driver_id = request.form.get('driver_id')
            if driver_id:
                with sqlite3.connect(DATABASE) as conn:
                    conn.execute("DELETE FROM drivers WHERE id = ?", (driver_id,))
                message = "Driver removed."
        elif action == 'add_vehicle':
            new_vehicle = request.form.get('new_vehicle', '').strip()
            new_vehicle_no = request.form.get('new_vehicle_no', '').strip()
            if new_vehicle:
                with sqlite3.connect(DATABASE) as conn:
                    try:
                        conn.execute("INSERT INTO vehicles (name, vehicle_no) VALUES (?, ?)", (new_vehicle, new_vehicle_no or None))
                        message = f"Vehicle '{new_vehicle}' added."
                    except sqlite3.IntegrityError:
                        message = f"Vehicle '{new_vehicle}' already exists."
                        message_type = 'error'
        elif action == 'delete_vehicle':
            vehicle_id = request.form.get('vehicle_id')
            if vehicle_id:
                with sqlite3.connect(DATABASE) as conn:
                    conn.execute("DELETE FROM vehicles WHERE id = ?", (vehicle_id,))
                message = "Vehicle removed."
        elif action == 'add_customer':
            new_customer = request.form.get('new_customer', '').strip()
            new_company  = request.form.get('new_company', '').strip()
            if new_customer:
                with sqlite3.connect(DATABASE) as conn:
                    try:
                        conn.execute("INSERT INTO customers (name, company) VALUES (?, ?)", (new_customer, new_company or None))
                        message = f"Customer '{new_customer}' added."
                    except sqlite3.IntegrityError:
                        message = f"Customer '{new_customer}' already exists."
                        message_type = 'error'
        elif action == 'delete_customer':
            customer_id = request.form.get('customer_id')
            if customer_id:
                with sqlite3.connect(DATABASE) as conn:
                    conn.execute("DELETE FROM customers WHERE id = ?", (customer_id,))
                message = "Customer removed."
        _cache_bust()

    # Batch all settings queries → 1 HTTP round-trip on Turso
    _sc = _db_multi_exec([
        ("SELECT COUNT(*) FROM invoices", ()),
        ("SELECT COUNT(*) FROM invoices WHERE strftime('%Y-%m', created_at) = strftime('%Y-%m','now')", ()),
        ("SELECT id, name FROM drivers ORDER BY name ASC", ()),
        ("SELECT id, name, COALESCE(vehicle_no,'') FROM vehicles ORDER BY name ASC", ()),
        ("SELECT id, name, COALESCE(company,'') FROM customers ORDER BY name ASC", ()),
        ("SELECT COUNT(*) FROM invoices WHERE COALESCE(bill_status,'Bill Generated') = 'Bill Generated'", ()),
        ("SELECT COUNT(*) FROM invoices WHERE bill_status = 'Bill Submitted'", ()),
        ("SELECT COUNT(*) FROM invoices WHERE bill_status = 'Payment Received'", ()),
        ("""SELECT vehicle_type, strftime('%Y-%m', date) as month,
                   ROUND(SUM(CAST(NULLIF(total_km,'') AS REAL)), 1) as km
            FROM invoices
            WHERE vehicle_type IS NOT NULL AND vehicle_type != ''
              AND date IS NOT NULL AND date != ''
            GROUP BY vehicle_type, month ORDER BY month DESC, vehicle_type ASC""", ()),
        ("""SELECT strftime('%Y-%m', date) AS month, COUNT(*) AS slips,
                   SUM(CASE WHEN COALESCE(bill_status,'Bill Generated')='Bill Generated' THEN 1 ELSE 0 END) AS not_submitted,
                   SUM(CASE WHEN bill_status='Bill Submitted' THEN 1 ELSE 0 END) AS submitted,
                   SUM(CASE WHEN bill_status='Payment Received' THEN 1 ELSE 0 END) AS paid,
                   ROUND(SUM(CAST(NULLIF(total_km,'') AS REAL)),1) AS km
            FROM invoices WHERE date IS NOT NULL AND date != ''
            GROUP BY month ORDER BY month DESC""", ()),
    ])
    total_invoices       = (_sc[0].fetchone() or (0,))[0]
    invoices_this_month  = (_sc[1].fetchone() or (0,))[0]
    drivers              = _sc[2].fetchall()
    vehicles             = _sc[3].fetchall()
    customers            = _sc[4].fetchall()
    not_submitted        = (_sc[5].fetchone() or (0,))[0]
    bill_submitted       = (_sc[6].fetchone() or (0,))[0]
    payment_received     = (_sc[7].fetchone() or (0,))[0]
    vehicle_km_rows      = _sc[8].fetchall()
    monthly_report_rows  = _sc[9].fetchall()

    return render_template('settings.html',
                           message=message,
                           message_type=message_type,
                           total_invoices=total_invoices,
                           invoices_this_month=invoices_this_month,
                           drivers=drivers,
                           vehicles=vehicles,
                           customers=customers,
                           vehicle_km_rows=vehicle_km_rows,
                           monthly_report_rows=monthly_report_rows,
                           not_submitted=not_submitted,
                           bill_submitted=bill_submitted,
                           payment_received=payment_received)


@app.route('/export/monthly_report')
def export_monthly_report():
    if 'admin' not in session:
        return redirect(url_for('login_page'))
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Accept ?months=2026-04,2026-03  OR  ?months=2026-04 (single) OR nothing (last 6)
    months_param = request.args.get('months', '').strip()
    month_list = [m.strip() for m in months_param.split(',') if m.strip()] if months_param else []

    with sqlite3.connect(DATABASE) as conn:
        if month_list:
            ph = ','.join('?' * len(month_list))
            summary_rows = conn.execute(f"""
                SELECT strftime('%Y-%m', date) AS m,
                       COUNT(*) AS slips,
                       SUM(CASE WHEN COALESCE(bill_status,'Bill Generated')='Bill Generated' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN bill_status='Bill Submitted' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN bill_status='Payment Received' THEN 1 ELSE 0 END),
                       ROUND(SUM(CAST(NULLIF(total_km,'') AS REAL)),1)
                FROM invoices WHERE date IS NOT NULL AND date != ''
                  AND strftime('%Y-%m', date) IN ({ph})
                GROUP BY m ORDER BY m DESC
            """, month_list).fetchall()
            km_rows = conn.execute(f"""
                SELECT vehicle_type, strftime('%Y-%m', date),
                       ROUND(SUM(CAST(NULLIF(total_km,'') AS REAL)),1)
                FROM invoices
                WHERE vehicle_type IS NOT NULL AND vehicle_type != ''
                  AND date IS NOT NULL AND date != ''
                  AND strftime('%Y-%m', date) IN ({ph})
                GROUP BY vehicle_type, strftime('%Y-%m', date)
                ORDER BY strftime('%Y-%m', date) DESC, vehicle_type ASC
            """, month_list).fetchall()
        else:
            summary_rows = conn.execute("""
                SELECT strftime('%Y-%m', date) AS m,
                       COUNT(*) AS slips,
                       SUM(CASE WHEN COALESCE(bill_status,'Bill Generated')='Bill Generated' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN bill_status='Bill Submitted' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN bill_status='Payment Received' THEN 1 ELSE 0 END),
                       ROUND(SUM(CAST(NULLIF(total_km,'') AS REAL)),1)
                FROM invoices WHERE date IS NOT NULL AND date != ''
                  AND strftime('%Y-%m', date) >= strftime('%Y-%m', date('now','-5 months'))
                GROUP BY m ORDER BY m DESC
            """).fetchall()
            km_rows = conn.execute("""
                SELECT vehicle_type, strftime('%Y-%m', date),
                       ROUND(SUM(CAST(NULLIF(total_km,'') AS REAL)),1)
                FROM invoices
                WHERE vehicle_type IS NOT NULL AND vehicle_type != ''
                  AND date IS NOT NULL AND date != ''
                  AND strftime('%Y-%m', date) >= strftime('%Y-%m', date('now','-5 months'))
                GROUP BY vehicle_type, strftime('%Y-%m', date)
                ORDER BY strftime('%Y-%m', date) DESC, vehicle_type ASC
            """).fetchall()

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1967D2")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style='thin', color='CCCCCC')
    cell_border = Border(left=thin, right=thin, bottom=thin)

    def _style_hdr(cell, text):
        cell.value = text
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # ── Sheet 1: Slip Summary ──
    ws1 = wb.active
    ws1.title = "Slip Summary"
    ws1.row_dimensions[1].height = 22
    for col, h in enumerate(["Month", "Total Slips", "Not Submitted", "Bill Submitted", "Payment Received", "Total KM"], 1):
        _style_hdr(ws1.cell(row=1, column=col), h)
    for r, row in enumerate(summary_rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws1.cell(row=r, column=c, value=val if val is not None else 0)
            cell.border = cell_border
            if c > 1:
                cell.alignment = Alignment(horizontal="center")
    _auto_width(ws1)

    # ── Sheet 2: Vehicle KM ──
    ws2 = wb.create_sheet("Vehicle KM")
    ws2.row_dimensions[1].height = 22
    for col, h in enumerate(["Vehicle", "Month", "Total KM (km)"], 1):
        _style_hdr(ws2.cell(row=1, column=col), h)
    for r, row in enumerate(km_rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val if val is not None else 0)
            cell.border = cell_border
            if c > 1:
                cell.alignment = Alignment(horizontal="center")
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    if month_list and len(month_list) == 1:
        stamp = month_list[0]
    elif month_list:
        stamp = f"{month_list[-1]}_to_{month_list[0]}"
    else:
        stamp = datetime.now().strftime('%Y-%m')
    return send_file(buf, download_name=f'slip_report_{stamp}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/templates', methods=['GET'])
def templates_list():
    if 'admin' not in session:
        return redirect(url_for('home'))
    admin_username = session['admin']
    # Pre-fill params from generator "Save as Template" link
    prefill = {k: request.args.get(k, '') for k in (
        'template_name', 'customer_name', 'company_name', 'driver_name',
        'vehicle_type', 'vehicle_no', 'route_covered', 'dn', 'remarks',
        'starting_km', 'total_km')}
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(
            """SELECT id, template_name, customer_name, company_name,
                      driver_name, vehicle_type, vehicle_no, route_covered,
                      created_at, use_count, last_used
               FROM slip_templates
               WHERE admin_username = ?
               ORDER BY use_count DESC, created_at DESC""",
            (admin_username,),
        ).fetchall()
    return render_template('templates_list.html', templates=rows, prefill=prefill)


@app.route('/templates/new')
def templates_new():
    if 'admin' not in session:
        return redirect(url_for('home'))
    driver_list, vehicle_rows, customer_rows = _load_ref_data()
    vehicle_list  = [r[0] for r in vehicle_rows]
    vehicle_map   = {r[0]: r[1] for r in vehicle_rows}
    customer_list = [r[0] for r in customer_rows]
    customer_map  = {r[0]: r[1] for r in customer_rows}
    # Pre-fill from generator "Save as Template" click
    prefill = {k: request.args.get(k, '') for k in (
        'customer_name', 'company_name', 'driver_name', 'vehicle_type',
        'vehicle_no', 'route_covered', 'dn', 'remarks', 'starting_km',
        'total_km', 'project_code', 'mail_approval_date')}
    return render_template('template_new.html',
                           driver_list=driver_list,
                           vehicle_list=vehicle_list, vehicle_map=vehicle_map,
                           customer_list=customer_list, customer_map=customer_map,
                           prefill=prefill)


@app.route('/templates/save', methods=['POST'])
def templates_save():
    if 'admin' not in session:
        return redirect(url_for('home'))
    admin_username = session['admin']
    name = (request.form.get('template_name') or '').strip()
    if not name:
        return redirect(url_for('templates_list'))
    fields = {k: (request.form.get(k) or '').strip() for k in (
        'customer_name', 'company_name', 'vehicle_type', 'vehicle_no',
        'route_covered', 'dn', 'remarks', 'driver_name',
        'starting_km', 'total_km', 'project_code', 'mail_approval_date',
        'starting_time', 'closing_time', 'route_stops_json')}
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """INSERT INTO slip_templates
               (admin_username, template_name, customer_name, company_name,
                vehicle_type, vehicle_no, route_covered, dn, remarks,
                driver_name, starting_km, total_km, project_code,
                mail_approval_date, starting_time, closing_time,
                route_stops_json, created_at, use_count)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0)""",
            (admin_username, name,
             fields['customer_name'], fields['company_name'],
             fields['vehicle_type'], fields['vehicle_no'],
             fields['route_covered'], fields['dn'], fields['remarks'],
             fields['driver_name'], fields['starting_km'], fields['total_km'],
             fields['project_code'], fields['mail_approval_date'],
             fields['starting_time'], fields['closing_time'],
             fields['route_stops_json'],
             datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        )
    return redirect(url_for('templates_list'))


@app.route('/templates/<int:template_id>/delete', methods=['POST'])
def templates_delete(template_id):
    if 'admin' not in session:
        return redirect(url_for('home'))
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            "DELETE FROM slip_templates WHERE id = ? AND admin_username = ?",
            (template_id, session['admin']),
        )
    return redirect(url_for('templates_list'))


@app.route('/templates/<int:template_id>/use', methods=['POST'])
def templates_use(template_id):
    if 'admin' not in session:
        return redirect(url_for('home'))
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """UPDATE slip_templates
               SET use_count = COALESCE(use_count, 0) + 1,
                   last_used = ?
               WHERE id = ? AND admin_username = ?""",
            (datetime.now().strftime('%Y-%m-%d %H:%M'), template_id, session['admin']),
        )
    return redirect(url_for('generator_form', template_id=template_id))


# --------------------------
# DIGITAL SIGNATURES
# --------------------------

@app.route('/request_signature', methods=['POST'])
def request_signature():
    if 'admin' not in session:
        return redirect(url_for('home'))
    invoice_ids = request.form.getlist('selected_invoices')
    if not invoice_ids:
        return redirect(url_for('admin_portal'))

    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            "SELECT customer_name FROM invoices WHERE id = ?", (invoice_ids[0],)
        ).fetchone()
    customer_name = row[0] if row else ''

    token = secrets.token_urlsafe(16)
    now = datetime.now()
    created_at = now.strftime('%Y-%m-%d %H:%M:%S')
    expires_at = (now + timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
    ids_json = _json.dumps([int(i) for i in invoice_ids])

    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """INSERT INTO signature_requests (token, invoice_ids, customer_name, created_at, expires_at)
               VALUES (?, ?, ?, ?, ?)""",
            (token, ids_json, customer_name, created_at, expires_at)
        )
        placeholders = ','.join('?' * len(invoice_ids))
        conn.execute(
            f"UPDATE invoices SET signature_status = 'pending' WHERE id IN ({placeholders}) AND COALESCE(signature_status,'') != 'signed'",
            invoice_ids
        )

    return redirect(url_for('signatures_page', new_token=token))


@app.route('/admin/signatures')
def signatures_page():
    if 'admin' not in session:
        return redirect(url_for('home'))
    new_token = request.args.get('new_token', '')
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(
            """SELECT id, token, invoice_ids, customer_name, created_at, expires_at, signed_at, signer_ip, signer_ua
               FROM signature_requests ORDER BY created_at DESC"""
        ).fetchall()

    now = datetime.now()
    now_str = now.strftime('%Y-%m-%d %H:%M:%S')
    requests_data = []
    for r in rows:
        ids = _json.loads(r[2]) if r[2] else []
        signed_at = r[6]
        expires_at = r[5]
        if signed_at:
            status = 'signed'
        elif expires_at and now_str > expires_at:
            status = 'expired'
        else:
            status = 'pending'
        # Compute hours since creation for reminder badge
        hours_pending = None
        if status == 'pending' and r[4]:
            try:
                created_dt = datetime.strptime(r[4], '%Y-%m-%d %H:%M:%S')
                hours_pending = int((now - created_dt).total_seconds() / 3600)
            except Exception:
                pass
        requests_data.append({
            'id': r[0], 'token': r[1], 'invoice_count': len(ids),
            'customer_name': r[3], 'created_at': r[4],
            'expires_at': expires_at, 'signed_at': signed_at,
            'signer_ip': r[7], 'signer_ua': r[8] or '',
            'device': _parse_ua(r[8] or ''),
            'hours_pending': hours_pending,
            'status': status,
        })

    return render_template('signatures.html', requests=requests_data, new_token=new_token)


@app.route('/admin/signatures/revoke/<int:req_id>', methods=['POST'])
def revoke_signature(req_id):
    if 'admin' not in session:
        return redirect(url_for('home'))
    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            "SELECT invoice_ids FROM signature_requests WHERE id = ?", (req_id,)
        ).fetchone()
        if row:
            ids = _json.loads(row[0]) if row[0] else []
            conn.execute("DELETE FROM signature_requests WHERE id = ?", (req_id,))
            if ids:
                # Find invoice IDs still covered by another active (unsigned) request
                remaining = conn.execute(
                    "SELECT invoice_ids FROM signature_requests WHERE signed_at IS NULL"
                ).fetchall()
                still_covered = set()
                for (r_ids,) in remaining:
                    try:
                        still_covered.update(_json.loads(r_ids))
                    except Exception:
                        pass
                for inv_id in ids:
                    if inv_id not in still_covered:
                        conn.execute(
                            "UPDATE invoices SET signature_status = NULL WHERE id = ? AND signature_status = 'pending'",
                            (inv_id,)
                        )
    return redirect(url_for('signatures_page'))


@app.route('/admin/signatures/reissue/<int:req_id>', methods=['POST'])
def reissue_signature(req_id):
    if 'admin' not in session:
        return redirect(url_for('home'))
    with sqlite3.connect(DATABASE) as conn:
        row = conn.execute(
            "SELECT invoice_ids, customer_name FROM signature_requests WHERE id = ?", (req_id,)
        ).fetchone()
        if not row:
            return redirect(url_for('signatures_page'))
        ids = _json.loads(row[0]) if row[0] else []
        customer_name = row[1]
        conn.execute("DELETE FROM signature_requests WHERE id = ?", (req_id,))
        # New token + 7-day window
        token = secrets.token_urlsafe(16)
        now = datetime.now()
        created_at = now.strftime('%Y-%m-%d %H:%M:%S')
        expires_at = (now + timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
        conn.execute(
            """INSERT INTO signature_requests (token, invoice_ids, customer_name, created_at, expires_at)
               VALUES (?, ?, ?, ?, ?)""",
            (token, _json.dumps(ids), customer_name, created_at, expires_at)
        )
        # Reset invoice status so the new request is active
        if ids:
            placeholders = ','.join('?' * len(ids))
            conn.execute(
                f"UPDATE invoices SET signature_status = 'pending', signed_at = NULL WHERE id IN ({placeholders})",
                ids
            )
    return redirect(url_for('signatures_page', new_token=token))


@app.route('/sign/<token>', methods=['GET'])
def sign_page(token):
    with sqlite3.connect(DATABASE) as conn:
        req = conn.execute(
            "SELECT id, invoice_ids, customer_name, created_at, expires_at, signed_at FROM signature_requests WHERE token = ?",
            (token,)
        ).fetchone()

    if not req:
        return render_template('sign.html', error='This signature link does not exist or has been revoked.')

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if req[5]:
        return render_template('sign.html', already_signed=True, signed_at=req[5], customer_name=req[2])
    if req[4] and now_str > req[4]:
        return render_template('sign.html', error='This signature link has expired. Please contact Osprey Travels for a new link.')

    ids = _json.loads(req[1]) if req[1] else []
    slips = []
    with sqlite3.connect(DATABASE) as conn:
        for inv_id in ids:
            row = conn.execute(
                """SELECT duty_slip_no, customer_name, company_name, date,
                          vehicle_type, vehicle_no, route_covered, driver_name
                   FROM invoices WHERE id = ?""",
                (inv_id,)
            ).fetchone()
            if row:
                slips.append({
                    'id': inv_id, 'duty_slip_no': row[0],
                    'customer_name': row[1], 'company_name': row[2],
                    'date': row[3], 'vehicle_type': row[4],
                    'vehicle_no': row[5], 'route_covered': row[6],
                    'driver_name': row[7],
                })

    return render_template('sign.html', token=token, slips=slips,
                           customer_name=req[2], expires_at=req[4])


@app.route('/sign/<token>', methods=['POST'])
def submit_signature(token):
    with sqlite3.connect(DATABASE) as conn:
        req = conn.execute(
            "SELECT id, invoice_ids, expires_at, signed_at FROM signature_requests WHERE token = ?",
            (token,)
        ).fetchone()
    if not req:
        return jsonify({'ok': False, 'error': 'Invalid link'})

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if req[3]:
        return jsonify({'ok': False, 'error': 'Already signed'})
    if req[2] and now_str > req[2]:
        return jsonify({'ok': False, 'error': 'Link expired'})

    sig_data = request.form.get('signature_data', '')
    if not sig_data or len(sig_data) < 100:
        return jsonify({'ok': False, 'error': 'Please draw your signature before submitting'})

    signer_ip = request.headers.get('X-Forwarded-For', request.remote_addr or '')
    signer_ua = request.headers.get('User-Agent', '')
    ids = _json.loads(req[1]) if req[1] else []

    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            "UPDATE signature_requests SET signed_at = ?, signature_data = ?, signer_ip = ?, signer_ua = ? WHERE id = ?",
            (now_str, sig_data, signer_ip, signer_ua, req[0])
        )
        if ids:
            placeholders = ','.join('?' * len(ids))
            conn.execute(
                f"UPDATE invoices SET signature_status = 'signed', signed_at = ? WHERE id IN ({placeholders})",
                [now_str] + ids
            )

    return jsonify({'ok': True})


# --------------------------
# PERFORMANCE: cache-busting on generate
# --------------------------
# generate_invoice already calls _cache_bust via settings; also bust on generate
# (the auto-create of customer/vehicle could add new entries)

# --------------------------
# DUPLICATE DETECTION
# --------------------------
@app.route('/check_duplicate', methods=['POST'])
def check_duplicate():
    if 'admin' not in session:
        return jsonify({'duplicates': []})
    body = request.get_json(silent=True) or {}
    customer_name = (body.get('customer_name') or request.form.get('customer_name') or '').strip()
    date_val      = (body.get('date') or request.form.get('date') or '').strip()
    vehicle_no    = (body.get('vehicle_no') or request.form.get('vehicle_no') or '').strip()
    if not customer_name or not date_val:
        return jsonify({'duplicates': []})
    with sqlite3.connect(DATABASE) as conn:
        params = [customer_name, date_val]
        extra = ''
        if vehicle_no:
            extra = " AND COALESCE(vehicle_no,'') = ?"
            params.append(vehicle_no)
        rows = conn.execute(
            f"SELECT id, duty_slip_no, date, route_covered, driver_name "
            f"FROM invoices WHERE customer_name = ? AND date = ?{extra} ORDER BY id DESC LIMIT 5",
            params
        ).fetchall()
    return jsonify({'duplicates': [
        {'id': r[0], 'duty_slip_no': r[1] or '—', 'date': r[2],
         'route': (r[3] or '')[:60], 'driver': r[4] or ''}
        for r in rows
    ]})


# --------------------------
# CUSTOMER HISTORY
# --------------------------
@app.route('/admin/customers')
def customers_page():
    if 'admin' not in session:
        return redirect(url_for('home'))
    new_portal = request.args.get('new_portal', '')
    portal_token = request.args.get('portal_token', '')
    merged = request.args.get('merged', '')
    add_error = request.args.get('add_error', '')
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute("""
            SELECT COALESCE(c.name, MIN(i.customer_name)) AS display_name,
                   COUNT(*) AS total,
                   MAX(i.date) AS last_date,
                   SUM(CASE WHEN i.signature_status = 'signed'  THEN 1 ELSE 0 END) AS signed_count,
                   SUM(CASE WHEN i.signature_status = 'pending' THEN 1 ELSE 0 END) AS pending_count,
                   SUM(CASE WHEN COALESCE(i.signature_status,'') = '' THEN 1 ELSE 0 END) AS unsigned_count,
                   MAX(c.portal_token) AS portal_token
            FROM invoices i
            LEFT JOIN customers c ON LOWER(c.name) = LOWER(i.customer_name)
            WHERE i.customer_name IS NOT NULL AND i.customer_name != ''
            GROUP BY LOWER(i.customer_name)
            ORDER BY MAX(i.date) DESC
        """).fetchall()
        all_customer_names = [r[0] for r in conn.execute(
            "SELECT COALESCE(c.name, MIN(i.customer_name)) FROM invoices i "
            "LEFT JOIN customers c ON LOWER(c.name) = LOWER(i.customer_name) "
            "WHERE i.customer_name IS NOT NULL AND i.customer_name != '' "
            "GROUP BY LOWER(i.customer_name) ORDER BY 1 ASC"
        ).fetchall()]
    customers_data = [
        {
            'name': r[0], 'total': r[1], 'last_date': r[2],
            'signed': r[3], 'pending': r[4], 'unsigned': r[5],
            'portal_token': r[6],
        }
        for r in rows
    ]
    portal_link = ''
    if new_portal and portal_token:
        portal_link = portal_token
    return render_template('customers.html', customers=customers_data,
                           all_names=all_customer_names,
                           new_portal=new_portal, portal_link=portal_link,
                           merged=merged, add_error=add_error)


@app.route('/admin/customers/portal_link', methods=['POST'])
def generate_portal_link():
    if 'admin' not in session:
        return redirect(url_for('home'))
    customer_name = (request.form.get('customer_name') or '').strip()
    if not customer_name:
        return redirect(url_for('customers_page'))
    token = secrets.token_urlsafe(24)
    with sqlite3.connect(DATABASE) as conn:
        exists = conn.execute("SELECT 1 FROM customers WHERE LOWER(name) = LOWER(?)", (customer_name,)).fetchone()
        if exists:
            conn.execute("UPDATE customers SET portal_token = ? WHERE LOWER(name) = LOWER(?)", (token, customer_name))
        else:
            conn.execute("INSERT INTO customers (name, portal_token) VALUES (?, ?)", (customer_name, token))
    return redirect(url_for('customers_page', new_portal=customer_name, portal_token=token))


# --------------------------
# CUSTOMER PORTAL (public)
# --------------------------
@app.route('/portal/<token>')
def customer_portal(token):
    with sqlite3.connect(DATABASE) as conn:
        cust = conn.execute(
            "SELECT name, COALESCE(company,'') FROM customers WHERE portal_token = ?", (token,)
        ).fetchone()
    if not cust:
        return render_template('portal.html', error=True)
    customer_name = cust[0]
    with sqlite3.connect(DATABASE) as conn:
        slips = conn.execute("""
            SELECT id, duty_slip_no, date, company_name, vehicle_type, vehicle_no,
                   driver_name, route_covered, total_km,
                   COALESCE(bill_status, 'Bill Generated'),
                   COALESCE(signature_status, '')
            FROM invoices WHERE customer_name = ?
            ORDER BY date DESC, id DESC
        """, (customer_name,)).fetchall()
    return render_template('portal.html',
                           customer_name=customer_name, company=cust[1],
                           slips=slips, token=token, error=False)


@app.route('/portal/<token>/slip/<int:invoice_id>')
def portal_slip_download(token, invoice_id):
    with sqlite3.connect(DATABASE) as conn:
        cust = conn.execute("SELECT name FROM customers WHERE portal_token = ?", (token,)).fetchone()
        if not cust:
            return "Invalid portal link", 403
        row = conn.execute(
            """SELECT customer_name, company_name, date, duty_slip_no, vehicle_type,
                      vehicle_no, starting_km, closing_km, total_km,
                      starting_time, closing_time, total_time,
                      project_code, mail_approval_date, route_covered, driver_name
               FROM invoices WHERE id = ? AND customer_name = ?""",
            (invoice_id, cust[0])
        ).fetchone()
        if not row:
            return "Slip not found", 404
        sig_data = _get_sig_for_invoice(conn, invoice_id)
    data = {
        'customer_name': row[0], 'company_name': row[1], 'date': row[2],
        'duty_slip_no': row[3], 'vehicle_type': row[4], 'vehicle_no': row[5],
        'starting_km': row[6], 'closing_km': row[7], 'total_km': row[8],
        'starting_time': row[9], 'closing_time': row[10], 'total_time': row[11],
        'project_code': row[12], 'mail_approval_date': row[13],
        'route_covered': row[14], 'driver_name': row[15],
    }
    buf = _build_pdf(data, signature_data=sig_data)
    filename = f"slip_{row[3] or row[0].replace(' ', '_')}.pdf"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


@app.route('/admin/signatures/sign_all_customer', methods=['POST'])
def sign_all_customer():
    if 'admin' not in session:
        return redirect(url_for('home'))
    customer_name = request.form.get('customer_name', '').strip()
    if not customer_name:
        return redirect(url_for('customers_page'))
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(
            """SELECT id FROM invoices
               WHERE customer_name = ? AND COALESCE(signature_status,'') != 'signed'""",
            (customer_name,)
        ).fetchall()
    ids = [r[0] for r in rows]
    if not ids:
        return redirect(url_for('customers_page'))
    token = secrets.token_urlsafe(16)
    now = datetime.now()
    created_at = now.strftime('%Y-%m-%d %H:%M:%S')
    expires_at = (now + timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
    with sqlite3.connect(DATABASE) as conn:
        conn.execute(
            """INSERT INTO signature_requests (token, invoice_ids, customer_name, created_at, expires_at)
               VALUES (?, ?, ?, ?, ?)""",
            (token, _json.dumps(ids), customer_name, created_at, expires_at)
        )
        placeholders = ','.join('?' * len(ids))
        conn.execute(
            f"UPDATE invoices SET signature_status = 'pending' WHERE id IN ({placeholders}) AND COALESCE(signature_status,'') != 'signed'",
            ids
        )
    return redirect(url_for('signatures_page', new_token=token))


@app.route('/admin/customers/add', methods=['POST'])
def add_customer():
    if 'admin' not in session:
        return redirect(url_for('home'))
    name = (request.form.get('name') or '').strip()
    company = (request.form.get('company') or '').strip()
    error = ''
    if name:
        with sqlite3.connect(DATABASE) as conn:
            exists = conn.execute("SELECT 1 FROM customers WHERE LOWER(name) = LOWER(?)", (name,)).fetchone()
            if exists:
                error = f'Customer "{name}" already exists.'
            else:
                conn.execute("INSERT INTO customers (name, company) VALUES (?, ?)", (name, company))
                _cache_bust()
    return redirect(url_for('customers_page', add_error=error) if error else url_for('customers_page'))


@app.route('/admin/customers/merge', methods=['POST'])
def merge_customers():
    if 'admin' not in session:
        return redirect(url_for('home'))
    from_name = (request.form.get('from_name') or '').strip()
    to_name   = (request.form.get('to_name') or '').strip()
    if from_name and to_name and from_name.lower() != to_name.lower():
        with sqlite3.connect(DATABASE) as conn:
            conn.execute("UPDATE invoices SET customer_name = ? WHERE LOWER(customer_name) = LOWER(?)", (to_name, from_name))
            conn.execute("UPDATE signature_requests SET customer_name = ? WHERE LOWER(customer_name) = LOWER(?)", (to_name, from_name))
            conn.execute("UPDATE slip_templates SET customer_name = ? WHERE LOWER(customer_name) = LOWER(?)", (to_name, from_name))
            conn.execute("DELETE FROM customers WHERE LOWER(name) = LOWER(?) AND LOWER(name) != LOWER(?)", (from_name, to_name))
        _cache_bust()
    return redirect(url_for('customers_page', merged='1'))


# --------------------------
# DRIVER REPORT
# --------------------------
@app.route('/admin/drivers')
def drivers_page():
    if 'admin' not in session:
        return redirect(url_for('home'))
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute("""
            SELECT driver_name,
                   COUNT(*) AS total,
                   MAX(date) AS last_date,
                   MIN(date) AS first_date,
                   ROUND(SUM(CAST(NULLIF(total_km,'') AS REAL)), 1) AS total_km
            FROM invoices
            WHERE driver_name IS NOT NULL AND driver_name != ''
            GROUP BY driver_name
            ORDER BY MAX(date) DESC
        """).fetchall()
    drivers_data = [
        {'name': r[0], 'total': r[1], 'last_date': r[2], 'first_date': r[3], 'total_km': r[4] or 0}
        for r in rows
    ]
    today_str = date.today().strftime('%Y-%m-%d')
    driver_list_all, _, _ = _load_ref_data()
    add_error = request.args.get('add_error', '')
    return render_template('drivers.html', drivers=drivers_data, today_str=today_str,
                           driver_list_all=driver_list_all, add_error=add_error)


@app.route('/admin/drivers/add', methods=['POST'])
def add_driver():
    if 'admin' not in session:
        return redirect(url_for('home'))
    name = (request.form.get('name') or '').strip()
    error = ''
    if name:
        with sqlite3.connect(DATABASE) as conn:
            try:
                conn.execute("INSERT INTO drivers (name) VALUES (?)", (name,))
                _cache_bust()
            except sqlite3.IntegrityError:
                error = f'Driver "{name}" already exists.'
    return redirect(url_for('drivers_page', add_error=error) if error else url_for('drivers_page'))


@app.route('/admin/drivers/<path:driver_name>/report')
def driver_report(driver_name):
    if 'admin' not in session:
        return redirect(url_for('login_page'))
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    where = "WHERE driver_name = ?"
    params = [driver_name]
    if date_from:
        where += " AND date >= ?"
        params.append(date_from)
    if date_to:
        where += " AND date <= ?"
        params.append(date_to)
    with sqlite3.connect(DATABASE) as conn:
        rows = conn.execute(
            f"""SELECT id, duty_slip_no, customer_name, company_name, date,
                       vehicle_type, vehicle_no, starting_km, closing_km, total_km,
                       starting_time, closing_time, total_time,
                       project_code, mail_approval_date, route_covered,
                       COALESCE(bill_status, 'Bill Generated')
                FROM invoices {where}
                ORDER BY date ASC, id ASC""",
            params
        ).fetchall()
        sig_map = _build_sig_map(conn, [r[0] for r in rows])
    if not rows:
        return "No slips found for this driver and date range.", 404
    slip_summaries = [
        {'duty_slip_no': r[1], 'customer_name': r[2], 'date': r[4],
         'route_covered': r[15], 'bill_status': r[16]}
        for r in rows
    ]
    date_range = ''
    if date_from or date_to:
        date_range = f'{_format_date(date_from) if date_from else "—"}  to  {_format_date(date_to) if date_to else "—"}'
    total_km = sum(
        float(r[9]) for r in rows if r[9] and str(r[9]).replace('.', '', 1).isdigit()
    )
    subtitle = f'{driver_name}   |   {date_range or "All time"}   |   {total_km:.1f} km total'
    pdf_libs = _pdf_libs()
    writer = pdf_libs['PdfWriter']()
    cover_reader = pdf_libs['PdfReader'](io.BytesIO(
        _build_cover_page('Driver Report', slip_summaries, subtitle).read()
    ))
    for page in cover_reader.pages:
        writer.add_page(page)
    for r in rows:
        data = {
            'duty_slip_no': r[1], 'customer_name': r[2], 'company_name': r[3],
            'date': r[4], 'vehicle_type': r[5], 'vehicle_no': r[6],
            'starting_km': r[7], 'closing_km': r[8], 'total_km': r[9],
            'starting_time': r[10], 'closing_time': r[11], 'total_time': r[12],
            'project_code': r[13], 'mail_approval_date': r[14],
            'route_covered': r[15], 'driver_name': driver_name,
        }
        slip_reader = pdf_libs['PdfReader'](io.BytesIO(
            _build_pdf(data, signature_data=sig_map.get(r[0])).read()
        ))
        for page in slip_reader.pages:
            writer.add_page(page)
    buf = io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    safe_name = driver_name.replace(' ', '_')
    stamp = f'_{date_from}_{date_to}' if (date_from or date_to) else ''
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'driver_report_{safe_name}{stamp}.pdf',
                     as_attachment=True)



# --------------------------
# SLIP MANAGEMENT
# --------------------------
def _slip_mgmt_query(year, month, status, date_type):
    col = "date" if date_type == 'duty' else "created_at"
    where = f"WHERE strftime('%Y', {col}) = ?"
    params = [str(year)]
    if month:
        where += f" AND strftime('%m', {col}) = ?"
        params.append(f"{int(month):02d}")
    STATUS_MAP = {
        'generated': "COALESCE(bill_status,'Bill Generated') = 'Bill Generated'",
        'submitted':  "bill_status = 'Bill Submitted'",
        'paid':       "bill_status = 'Payment Received'",
        'signed':     "signature_status = 'signed'",
    }
    if status and status in STATUS_MAP:
        where += f" AND {STATUS_MAP[status]}"
    return where, params


@app.route('/admin/slips')
def slip_management():
    if 'admin' not in session:
        return redirect(url_for('home'))
    year      = int(request.args.get('year', date.today().year))
    month     = request.args.get('month', '')
    status    = request.args.get('status', 'all')
    date_type = request.args.get('date_type', 'duty')
    col       = "date" if date_type == 'duty' else "created_at"

    with sqlite3.connect(DATABASE) as conn:
        # Available years
        years = [r[0] for r in conn.execute(
            "SELECT DISTINCT strftime('%Y', date) FROM invoices WHERE date IS NOT NULL ORDER BY 1 DESC"
        ).fetchall() if r[0]]
        if str(year) not in years and years:
            years.insert(0, str(year))
        years = sorted(set(years), reverse=True)

        # Month counts for sidebar
        where_y, params_y = _slip_mgmt_query(year, None, status, date_type)
        month_rows = conn.execute(
            f"SELECT strftime('%m', {col}) AS m, COUNT(*) FROM invoices {where_y} GROUP BY m",
            params_y
        ).fetchall()
        month_counts = {int(r[0]): r[1] for r in month_rows if r[0]}

        # Status tab counts (for selected year + month)
        where_base, params_base = _slip_mgmt_query(year, month, 'all', date_type)
        status_counts = {}
        total = conn.execute(f"SELECT COUNT(*) FROM invoices {where_base}", params_base).fetchone()[0]
        status_counts['all'] = total
        for s in ('generated', 'submitted', 'paid', 'signed'):
            w, p = _slip_mgmt_query(year, month, s, date_type)
            status_counts[s] = conn.execute(f"SELECT COUNT(*) FROM invoices {w}", p).fetchone()[0]

        # Slips for current view
        where, params = _slip_mgmt_query(year, month, status, date_type)
        slips = conn.execute(
            f"""SELECT id, duty_slip_no, date, created_at, customer_name, company_name,
                       route_covered, driver_name, vehicle_type, vehicle_no,
                       COALESCE(bill_status,'Bill Generated'), COALESCE(signature_status,'')
                FROM invoices {where} ORDER BY {col} DESC, id DESC LIMIT 500""",
            params
        ).fetchall()

    return render_template('slip_management.html',
                           slips=slips, year=year, years=years, month=int(month) if month else 0,
                           status=status, date_type=date_type,
                           month_counts=month_counts, status_counts=status_counts)


@app.route('/admin/slips/rows')
def slip_management_rows():
    if 'admin' not in session:
        return '', 401
    year      = int(request.args.get('year', date.today().year))
    month     = request.args.get('month', '')
    status    = request.args.get('status', 'all')
    date_type = request.args.get('date_type', 'duty')
    col       = "date" if date_type == 'duty' else "created_at"
    where, params = _slip_mgmt_query(year, month, status, date_type)
    with sqlite3.connect(DATABASE) as conn:
        slips = conn.execute(
            f"""SELECT id, duty_slip_no, date, created_at, customer_name, company_name,
                       route_covered, driver_name, vehicle_type, vehicle_no,
                       COALESCE(bill_status,'Bill Generated'), COALESCE(signature_status,'')
                FROM invoices {where} ORDER BY {col} DESC, id DESC LIMIT 500""",
            params
        ).fetchall()
    return render_template('_slip_management_rows.html', slips=slips, status=status)


@app.route('/admin/slips/month_counts')
def slip_management_month_counts():
    if 'admin' not in session:
        return jsonify({}), 401
    year      = int(request.args.get('year', date.today().year))
    status    = request.args.get('status', 'all')
    date_type = request.args.get('date_type', 'duty')
    col       = "date" if date_type == 'duty' else "created_at"
    where_y, params_y = _slip_mgmt_query(year, None, status, date_type)
    with sqlite3.connect(DATABASE) as conn:
        month_rows = conn.execute(
            f"SELECT strftime('%m', {col}) AS m, COUNT(*) FROM invoices {where_y} GROUP BY m",
            params_y
        ).fetchall()
    return jsonify({int(r[0]): r[1] for r in month_rows if r[0]})


# --------------------------
# BULK IMPORT
# --------------------------
@app.route('/admin/bulk_import', methods=['GET'])
def bulk_import_page():
    if 'admin' not in session:
        return redirect(url_for('home'))
    return render_template('bulk_import.html')


@app.route('/admin/bulk_import/parse', methods=['POST'])
def bulk_import_parse():
    if 'admin' not in session:
        return jsonify({'ok': False, 'error': 'Not authenticated'})
    f = request.files.get('file')
    if not f:
        return jsonify({'ok': False, 'error': 'No file uploaded'})
    filename = (f.filename or '').lower()
    rows = []
    try:
        if filename.endswith('.csv'):
            import csv as _csv
            content = f.read().decode('utf-8-sig')
            reader = _csv.DictReader(content.splitlines())
            for row in reader:
                rows.append({(k or '').strip().lower().replace(' ', '_'): (v or '').strip() for k, v in row.items()})
        elif filename.endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            ws = wb.active
            headers = None
            for r in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c or '').strip().lower().replace(' ', '_') for c in r]
                else:
                    rows.append({headers[i]: str(r[i] or '').strip() for i in range(min(len(headers), len(r)))})
            wb.close()
        else:
            return jsonify({'ok': False, 'error': 'Use a .csv or .xlsx file'})
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Parse error: {e}'})

    ALIASES = {
        'slip_no': 'duty_slip_no', 'slip#': 'duty_slip_no', 'slip_number': 'duty_slip_no',
        'customer': 'customer_name', 'company': 'company_name',
        'vehicle': 'vehicle_type', 'vehicle_number': 'vehicle_no',
        'driver': 'driver_name', 'route': 'route_covered',
        'start_km': 'starting_km', 'end_km': 'closing_km',
        'start_time': 'starting_time', 'end_time': 'closing_time',
        'project': 'project_code',
    }
    FIELDS = {'duty_slip_no', 'customer_name', 'company_name', 'date', 'vehicle_type',
              'vehicle_no', 'driver_name', 'route_covered', 'starting_km', 'closing_km',
              'total_km', 'starting_time', 'closing_time', 'project_code'}

    parsed, errors = [], []
    for i, row in enumerate(rows, 1):
        norm = {}
        for k, v in row.items():
            canonical = ALIASES.get(k, k)
            if canonical in FIELDS:
                norm[canonical] = v
        if not norm.get('customer_name') or not norm.get('date'):
            errors.append(f'Row {i}: missing customer_name or date — skipped')
            continue
        parsed.append(norm)
    return jsonify({'ok': True, 'rows': parsed, 'errors': errors, 'total': len(parsed)})


@app.route('/admin/bulk_import/generate', methods=['POST'])
def bulk_import_generate():
    if 'admin' not in session:
        return redirect(url_for('home'))
    try:
        rows = _json.loads(request.form.get('rows_json', '[]'))
    except Exception:
        return redirect(url_for('bulk_import_page'))
    if not rows:
        return redirect(url_for('bulk_import_page'))

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    slip_data_list = []
    with sqlite3.connect(DATABASE) as conn:
        for row in rows:
            sno = row.get('duty_slip_no') or get_next_duty_slip_no()
            data = {
                'customer_name':    row.get('customer_name', ''),
                'company_name':     row.get('company_name', ''),
                'date':             row.get('date', ''),
                'duty_slip_no':     sno,
                'vehicle_type':     row.get('vehicle_type', ''),
                'vehicle_no':       row.get('vehicle_no', ''),
                'starting_km':      row.get('starting_km', ''),
                'closing_km':       row.get('closing_km', ''),
                'total_km':         row.get('total_km', ''),
                'starting_time':    row.get('starting_time', ''),
                'closing_time':     row.get('closing_time', ''),
                'total_time':       row.get('total_time', ''),
                'project_code':     row.get('project_code', ''),
                'mail_approval_date': '',
                'route_covered':    row.get('route_covered', ''),
                'driver_name':      row.get('driver_name', ''),
            }
            conn.execute("""
                INSERT INTO invoices
                (customer_name, company_name, date, duty_slip_no,
                 vehicle_type, vehicle_no, starting_km, closing_km, total_km,
                 starting_time, closing_time, total_time, project_code,
                 route_covered, driver_name, admin_username, created_at, bill_status)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'Bill Generated')
            """, (data['customer_name'], data['company_name'], data['date'],
                  data['duty_slip_no'], data['vehicle_type'], data['vehicle_no'],
                  data['starting_km'], data['closing_km'], data['total_km'],
                  data['starting_time'], data['closing_time'], data['total_time'],
                  data['project_code'], data['route_covered'], data['driver_name'],
                  session['admin'], now_str))
            slip_data_list.append(data)
    _cache_bust()

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w') as zipf:
        for data in slip_data_list:
            pdf_buf = _build_pdf(data)
            fname = f"slip_{data['duty_slip_no'] or data['customer_name'].replace(' ', '_')}.pdf"
            zipf.writestr(fname, pdf_buf.read())
    zip_buf.seek(0)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(zip_buf, as_attachment=True,
                     download_name=f'bulk_slips_{stamp}.zip',
                     mimetype='application/zip')


@app.errorhandler(404)
def not_found(e):
    return ('''<!doctype html><html><head><title>404 — Osprey Travels</title>
<style>body{font-family:-apple-system,sans-serif;background:#f2f2f7;display:flex;align-items:center;
justify-content:center;min-height:100vh;margin:0;}
.box{text-align:center;padding:40px;background:#fff;border-radius:16px;box-shadow:0 2px 8px rgba(0,0,0,.08);}
h1{font-size:56px;margin:0;color:#1c1c1e;}p{color:#8e8e93;margin:8px 0 24px;}
a{background:#007aff;color:#fff;padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:600;}
</style></head><body><div class="box"><h1>404</h1><p>This page doesn't exist.</p>
<a href="/">Go Home</a></div></body></html>''', 404)


@app.errorhandler(500)
def server_error(e):
    return ('''<!doctype html><html><head><title>500 — Osprey Travels</title>
<style>body{font-family:-apple-system,sans-serif;background:#f2f2f7;display:flex;align-items:center;
justify-content:center;min-height:100vh;margin:0;}
.box{text-align:center;padding:40px;background:#fff;border-radius:16px;box-shadow:0 2px 8px rgba(0,0,0,.08);}
h1{font-size:56px;margin:0;color:#ff3b30;}p{color:#8e8e93;margin:8px 0 24px;}
a{background:#007aff;color:#fff;padding:10px 22px;border-radius:8px;text-decoration:none;font-weight:600;}
</style></head><body><div class="box"><h1>500</h1><p>Something went wrong on our end.</p>
<a href="/">Go Home</a></div></body></html>''', 500)


if __name__ == '__main__':
    app.run(debug=os.environ.get('FLASK_DEBUG', '0') == '1', port=5050)
